import logging
import os
from typing import List, Optional

import openpyxl
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import pyqtSlot as Slot

from .outliertest import OutlierTestWindow
from .processingwindow import ProcessingWindow
from .results_ui import Ui_Form
from .showanisotropy import ShowAnisotropyWindow
from .showcurve import ShowCurveWindow
from .showimage import ShowImageWindow
from .transmission import TransmissionWindow
from .vacuum import VacuumWindow
from ..utils.filebrowsers import getDirectory, getSaveFile
from ...core2.processing.calculations.resultsentry import SampleDistanceEntry, CurveFileType, PatternFileType, \
    CorMatFileType, SampleDistanceEntryType

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


class ResultsWindow(ProcessingWindow, Ui_Form):
    def setupUi(self, Form):
        super().setupUi(Form)
        model = QtCore.QSortFilterProxyModel()
        model.setSourceModel(self.project.results)
        self.treeView.setModel(model)
        self.treeView.model().modelReset.connect(self.resizeColumns)
        self.treeView.sortByColumn(0, QtCore.Qt.SortOrder.AscendingOrder)
        self.reloadPushButton.clicked.connect(self.project.results.reload)
        self.outlierTestPushButton.clicked.connect(self.openOutlierTest)
        self.project.results.modelReset.connect(self.resizeColumns)
        self.showCurvePushButton.clicked.connect(self.showCurve)
        self.show2DPushButton.clicked.connect(self.showImage)
        self.anisotropyPushButton.clicked.connect(self.showAnisotropy)
        for action in [self.actionASCII_txt, self.actionATSAS, self.actionRSR, self.actionExcel, self.actionPDH]:
            self.exportCurvesToolButton.addAction(action)
        #            action.triggered.connect(self.exportCurves)
        self.exportCurvesToolButton.setDefaultAction(self.actionASCII_txt)
        self.exportCurvesToolButton.triggered.connect(self.exportCurves)
        for action in [self.actionCmatNumpy_NPZ, self.actionCmatMatlab, self.actionCmatASCII, self.actionCmatAsciiGZip]:
            self.exportOutliersToolButton.addAction(action)
        #            action.triggered.connect(self.exportOutlierTestResults)
        self.exportOutliersToolButton.setDefaultAction(self.actionCmatNumpy_NPZ)
        self.exportOutliersToolButton.triggered.connect(self.exportOutlierTestResults)
        for action in [self.actionPatternNumpy_NPZ, self.actionPatternMatlab, self.actionPatternASCII,
                       self.actionPatternAsciiGZip]:
            self.exportPatternsToolButton.addAction(action)
        #            action.triggered.connect(self.exportPatterns)
        self.exportPatternsToolButton.setDefaultAction(self.actionPatternNumpy_NPZ)
        self.exportPatternsToolButton.triggered.connect(self.exportPatterns)
        self.timeBudgetPushButton.clicked.connect(self.showTimeBudget)
        self.transmissionPushButton.clicked.connect(self.showTransmission)
        self.vacuumFluxPushButton.clicked.connect(self.showVacuum)
        self.exportReportPushButton.clicked.connect(self.exportReport)
        self.cleanResultsPushButton.clicked.connect(self.cleanResults)
        self.removeSelectedResultPushButton.clicked.connect(self.removeSelectedResult)
        self.resizeColumns()

    @Slot()
    def cleanResults(self):
        self.project.settings.h5io.clear()
        self.project.results.reload()

    @Slot()
    def removeSelectedResult(self):
        sdes: List[SampleDistanceEntry] = [index.data(QtCore.Qt.ItemDataRole.UserRole) for index in
                                           self.treeView.selectionModel().selectedRows(0)]
        samples_dists = {(sde.samplename, sde.distancekey) for sde in sdes}
        for samplename, distkey in samples_dists:
            self.project.settings.h5io.removeDistance(samplename, distkey)
        self.project.settings.h5io.pruneSamplesWithNoDistances()
        self.project.results.reload()

    @Slot()
    def openOutlierTest(self):
        for index in self.treeView.selectionModel().selectedRows(0):
            sde = index.data(QtCore.Qt.ItemDataRole.UserRole)
            assert isinstance(sde, SampleDistanceEntry)
            if sde.entrytype != SampleDistanceEntryType.Primary:
                # only primary results have correlation matrices
                continue
            self.mainwindow.createViewWindow(OutlierTestWindow, [(sde.samplename, sde.distancekey)])

    @Slot()
    def resizeColumns(self):
        for c in range(self.treeView.model().columnCount()):
            self.treeView.resizeColumnToContents(c)

    @Slot()
    def showAnisotropy(self):
        logger.debug('showAnisotropy')
        for index in self.treeView.selectionModel().selectedRows(0):
            logger.debug('+')
            sde = index.data(QtCore.Qt.ItemDataRole.UserRole)
            if sde.entrytype not in [SampleDistanceEntryType.Subtracted, SampleDistanceEntryType.Primary]:
                continue
            self.mainwindow.createViewWindow(ShowAnisotropyWindow, [(sde.samplename, sde.distancekey)])

    @Slot()
    def showImage(self):
        for index in self.treeView.selectionModel().selectedRows(0):
            sde = index.data(QtCore.Qt.ItemDataRole.UserRole)
            if sde.entrytype not in [SampleDistanceEntryType.Subtracted, SampleDistanceEntryType.Primary]:
                continue
            self.mainwindow.createViewWindow(ShowImageWindow, [(sde.samplename, sde.distancekey)])

    @Slot()
    def showCurve(self):
        items = []
        for index in self.treeView.selectionModel().selectedRows(0):
            sde = index.data(QtCore.Qt.ItemDataRole.UserRole)
            items.append((sde.samplename, sde.distancekey))
        items = list(sorted(items))
        if items:
            self.mainwindow.createViewWindow(ShowCurveWindow, items)

    @Slot()
    def showTransmission(self):
        items = []
        for index in self.treeView.selectionModel().selectedRows(0):
            sde = index.data(QtCore.Qt.ItemDataRole.UserRole)
            assert isinstance(sde, SampleDistanceEntry)
            if sde.entrytype not in [SampleDistanceEntryType.Primary]:
                continue
            items.append((sde.samplename, sde.distancekey))
        items = list(sorted(items))
        if items:
            self.mainwindow.createViewWindow(TransmissionWindow, items)

    @Slot()
    def showTimeBudget(self):
        pass

    @Slot()
    def showVacuum(self):
        items = []
        for index in self.treeView.selectionModel().selectedRows(0):
            sde = index.data(QtCore.Qt.ItemDataRole.UserRole)
            assert isinstance(sde, SampleDistanceEntry)
            if sde.entrytype not in [SampleDistanceEntryType.Primary]:
                continue
            items.append((sde.samplename, sde.distancekey))
            logger.debug(f'Items now {items}')
        items = list(sorted(items))
        logger.debug(f'Sorted items: {items}')
        if items:
            self.mainwindow.createViewWindow(VacuumWindow, items)

    @Slot(QtWidgets.QAction)
    def exportPatterns(self, action: Optional[QtWidgets.QAction] = None):
        if action is None:
            action = self.sender()
        if not isinstance(action, QtWidgets.QAction):
            action = self.exportPatternsToolButton.defaultAction()
        logger.debug(f'Export pattern action: {action.objectName()}')
        self.exportPatternsToolButton.setDefaultAction(action)
        items: List[SampleDistanceEntry] = [index.data(QtCore.Qt.ItemDataRole.UserRole) for index in
                                            self.treeView.selectionModel().selectedRows(0)]
        items = [it for it in items if
                 it.entrytype in [SampleDistanceEntryType.Primary, SampleDistanceEntryType.Subtracted]]
        if not items:
            return
        outputfolder = getDirectory(self, 'Write scattering patterns to directory')
        if not outputfolder:
            return
        for item in items:
            basename = f'{item.samplename}_{item.distancekey}'
            filetype = {self.actionPatternASCII: PatternFileType.ASCII,
                        self.actionPatternAsciiGZip: PatternFileType.ASCIIGZIP,
                        self.actionPatternMatlab: PatternFileType.MATLAB,
                        self.actionPatternNumpy_NPZ: PatternFileType.NUMPY}[action]
            item.writePattern(os.path.join(outputfolder, basename + filetype.value[-1]), filetype)

    @Slot(QtWidgets.QAction)
    def exportOutlierTestResults(self, action: Optional[QtWidgets.QAction] = None):
        if action is None:
            action = self.sender()
        if not isinstance(action, QtWidgets.QAction):
            logger.debug(f'sender() is not an Action. It is a {type(action)}')
            action = self.exportOutliersToolButton.defaultAction()
        logger.debug(f'Outlier test export action: {action.objectName()}')
        self.exportOutliersToolButton.setDefaultAction(action)
        items: List[SampleDistanceEntry] = [index.data(QtCore.Qt.ItemDataRole.UserRole) for index in
                                            self.treeView.selectionModel().selectedRows(0)]
        items = [it for it in items if it.entrytype == SampleDistanceEntryType.Primary]
        if not items:
            return
        outputfolder = getDirectory(self, 'Write correlation matrices to directory:')
        if not outputfolder:
            return
        for item in items:
            basename = f'cmat_{item.samplename}_{item.distancekey}'
            filetype = {self.actionCmatASCII: CorMatFileType.ASCII,
                        self.actionCmatAsciiGZip: CorMatFileType.ASCIIGZIP,
                        self.actionCmatMatlab: CorMatFileType.MATLAB,
                        self.actionCmatNumpy_NPZ: CorMatFileType.NUMPY}[action]
            item.writeCorMat(os.path.join(outputfolder, basename + filetype.value[-1]), filetype)

    @Slot(QtWidgets.QAction)
    def exportCurves(self, action: Optional[QtWidgets.QAction] = None):
        if action is None:
            action = self.sender()
        if not isinstance(action, QtWidgets.QAction):
            action = self.exportCurvesToolButton.defaultAction()
        self.exportCurvesToolButton.setDefaultAction(action)
        items: List[SampleDistanceEntry] = [index.data(QtCore.Qt.ItemDataRole.UserRole) for index in
                                            self.treeView.selectionModel().selectedRows(0)]
        if not items:
            return
        if action == self.actionExcel:
            xlsxfile = getSaveFile(
                self, 'Select Excel Workbook file name...', '',
                'Microsoft(R) Excel(TM) Workbook files (*.xlsx);;All files (*)', '.xlsx')
            if not xlsxfile:
                return
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            wsnamecounter = {}
            for item in items:
                wsname = f'{item.samplename}_{item.distancekey}'
                if len(wsname) >= 31:
                    # some programs cannot open workbooks with too long sheet names
                    wsname = wsname[:29]
                    if wsname in wsnamecounter:
                        wsname = wsname + str(wsnamecounter[wsname])
                        wsnamecounter[wsname] += 1
                    else:
                        wsnamecounter[wsname] = 1
                ws = wb.create_sheet(wsname)
                item.writeCurveToXLSX(ws, ws.cell(row=1, column=1))
            wb.save(xlsxfile)
        else:
            outputfolder = getDirectory(self, 'Write scattering curves to directory:')
            if not outputfolder:
                return
            for item in items:
                basename = f'{item.samplename}_{item.distancekey}'
                filetype = {self.actionASCII_txt: CurveFileType.ASCII,
                            self.actionATSAS: CurveFileType.ATSAS,
                            self.actionPDH: CurveFileType.PDH,
                            self.actionRSR: CurveFileType.RSR}[action]
                item.writeCurve(os.path.join(outputfolder, basename + filetype.value[-1]), filetype)

    @Slot()
    def exportReport(self):
        pass
