import enum
from typing import Dict, Optional
import gzip
import logging

import openpyxl.worksheet.worksheet, openpyxl.cell
import scipy.io

from ...dataclasses import Exposure, Curve, Header, Sample
from .outliertest import OutlierTest
import weakref
import numpy as np

logger=logging.getLogger(__name__)
logger.setLevel(logging.INFO)


class CurveFileType(enum.Enum):
    ASCII = ('ASCII text file', '.txt')
    ATSAS = ('ATSAS data file', '.dat')
    PDH = ('PCG data file', '.pdh')
    RSR = ('RSR data file', '.rsr')

class PatternFileType(enum.Enum):
    ASCII = ('ASCII text file', '.txt')
    ASCIIGZIP = ('GZip-ped ASCII file', '.txt.gz')
    NUMPY = ('NumPy compressed', '.npz')
    MATLAB = ('MatLab(R) matrix', '.mat')

CorMatFileType = PatternFileType


class SampleDistanceEntryType(enum.Enum):
    Primary = 'primary'
    Subtracted = 'subtracted'
    Merged = 'merged'


class SampleDistanceEntry:
    """An entry in the HDF5 file corresponding one sample measured at one sample-to-detector distance"""
    samplename: str
    distancekey: str
    _entrytype: Optional[SampleDistanceEntryType] = None
    processing: "Processing"
    _header: Optional[Header] = None
    _exposure: Optional[Exposure] = None
    _outliertest:  Optional[OutlierTest] = None
    _curve: Optional[Curve] = None

    def __init__(self, samplename: str, distancekey: str, processing: "Processing"):
        self.samplename = samplename
        self.distancekey = distancekey
        try:
            self.processing = weakref.proxy(processing)
        except TypeError:
            # already a weakref proxy
            self.processing = processing


    def curves(self, all: bool=False) -> Dict[int, Curve]:
        return self.processing.settings.h5io.readCurves(self.h5path, readall=all)

    def headers(self, all: bool=False) -> Dict[int, Header]:
        return self.processing.settings.h5io.readHeaders(self.h5path, readall=all)


    @property
    def h5path(self) -> str:
        return f'Samples/{self.samplename}/{self.distancekey}'

    @property
    def curve(self) -> Curve:
        if self._curve is None:
            self._curve = self.processing.settings.h5io.readCurve(f'{self.h5path}/curve')
        return self._curve

    @property
    def exposure(self) -> Exposure:
        if self._exposure is None:
            self._exposure = self.processing.settings.h5io.readExposure(self.h5path)
        return self._exposure

    @property
    def outliertest(self) -> OutlierTest:
        if self.entrytype in [SampleDistanceEntryType.Merged, SampleDistanceEntryType.Subtracted]:
            return None
        if self._outliertest is None:
            self._outliertest = self.processing.settings.h5io.readOutlierTest(self.h5path)
        return self._outliertest

    @property
    def header(self) -> Header:
        if self._header is None:
            self._header = self.processing.settings.h5io.readHeader(self.h5path)
        return self._header

    @property
    def entrytype(self) -> SampleDistanceEntryType:
        if self._entrytype is None:
            cat = self.header.sample_category
            if cat == Sample.Categories.Subtracted.value:
                self._entrytype = SampleDistanceEntryType.Subtracted
            elif cat == Sample.Categories.Merged.value:
                self._entrytype = SampleDistanceEntryType.Merged
            else:
                self._entrytype = SampleDistanceEntryType.Primary
            logger.debug(f'Category: {cat=} -> {self._entrytype=}')
        return self._entrytype

    def writeCurve(self, filename: str, filetype: CurveFileType):
        with open(filename, 'wt') as f:
            dic = self.processing.settings.h5io.readHeaderDict(self.h5path)
            curvearray = np.array(self.curve, copy=True)
            if filetype in [CurveFileType.RSR, CurveFileType.ATSAS, CurveFileType.PDH]:
                curvearray[:,0]/=10
                curvearray[:,3]/=10
            if filetype == CurveFileType.RSR:
                f.write(' TIME\n')
                f.write(' 1.0\n')
                f.write(f' {curvearray.shape[0]}\n')
                np.savetxt(f, curvearray[:, :3], delimiter=' ', fmt='%.9f')
            elif filetype in [CurveFileType.ATSAS, CurveFileType.ASCII]:
                for key in sorted(dic):
                    f.write(f'# {key} : {dic[key]}\n')
                f.write('# Columns:\n')
                f.write('#  q [1/nm],  Intensity [1/cm * 1/sr], '
                        'Propagated uncertainty of the intensity [1/cm * 1/sr], '
                        'Propagated uncertainty of q [1/nm], Bin area [# of pixels], Pixel coordinate [pixel]\n'
                        )
                np.savetxt(f, curvearray if filetype == CurveFileType.ASCII else curvearray[:,:3])
            elif filetype == CurveFileType.PDH:
                f.write(f'{self.samplename} @ {self.distancekey} mm\n')
                f.write('SAXS\n')
                f.write(
                    f'{curvearray.shape[0]:>9d}         0         0         0         0         0         0         0\n'
                    '  0.000000E+00   0.000000E+00   0.000000E+00   1.000000E+00   0.000000E+00\n'
                    '  0.000000E+00   0.000000E+00   0.000000E+00   0.000000E+00   0.000000E+00\n'
                )
                np.savetxt(f, curvearray[:, :3], delimiter=' ', fmt='%14.6E')
            else:
                raise ValueError(f'Unknown curve file type: {filetype}')

    def writeCurveToXLSX(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, topleftcell: openpyxl.cell.Cell):
        row = topleftcell.row
        column = topleftcell.column
        worksheet.cell(row=row, column=column, value=f'{self.samplename} @ {self.distancekey} mm')
        worksheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column+3)
        for i, (quantity, unit) in enumerate([
            ('q', '1/nm'),
            ('Intensity', '1/nm'),
            ('dIntensity', '1/cm * 1/sr'),
            ('dq', '1/nm'),
        ]):
            worksheet.cell(row=row+1, column=column+i, value=quantity)
            worksheet.cell(row=row+2, column=column+i, value=unit)
        curvearray = np.array(self.curve)
        for r in range(len(self.curve)):
            for c in range(4):
                worksheet.cell(row=row+3+r, column=column+c, value=curvearray[r, c])

    def writePattern(self, filename: str, filetype: PatternFileType):
        if filetype == PatternFileType.NUMPY:
            np.savez_compressed(
                filename, intensity=self.exposure.intensity, error=self.exposure.uncertainty, mask=self.exposure.mask)
        elif filetype == PatternFileType.MATLAB:
            scipy.io.savemat(filename, {
                'intensity': self.exposure.intensity,
                'error': self.exposure.uncertainty,
                'mask': self.exposure.mask
            }, do_compression=True)
        elif filetype in [PatternFileType.ASCII, PatternFileType.ASCIIGZIP]:
            with (open(filename, 'wb') if (
                    filetype == PatternFileType.ASCII) else gzip.GzipFile(filename, 'w')) as f:
                f.write('# Intensity\n'.encode('utf-8'))
                np.savetxt(f, self.exposure.intensity)
                f.write('\n# Uncertainty\n'.encode('utf-8'))
                np.savetxt(f, self.exposure.uncertainty)
                f.write('# Mask\n'.encode('utf-8'))
                np.savetxt(f, self.exposure.mask)

    def writeCorMat(self, filename: str, filetype: CorMatFileType):
        if self.entrytype != SampleDistanceEntryType.Primary:
            raise ValueError('Cannot export correlation matrix of derived data.')
        if filetype == CorMatFileType.NUMPY:
            np.savez_compressed(
                filename, correlmatrix=self.outliertest.correlmatrix)
        elif filetype == CorMatFileType.MATLAB:
            scipy.io.savemat(filename, {
                'correlmatrix': self.outliertest.correlmatrix,
            }, do_compression=True)
        elif filetype in [CorMatFileType.ASCII, CorMatFileType.ASCIIGZIP]:
            with (open(filename, 'wt') if (
                    filetype == CorMatFileType.ASCII) else gzip.GzipFile(filename, 'w')) as f:
                np.savetxt(f, self.outliertest.correlmatrix)

    def isDerived(self) -> bool:
        return self.entrytype != SampleDistanceEntryType.Primary

