import configparser
import gc
import inspect
import logging
import os
import pickle
import queue
import string
import threading
from typing import Iterable, Optional, Union

import appdirs
import h5py
import matplotlib
import matplotlib.colors
import numpy as np
import openpyxl
import openpyxl.chart
import openpyxl.chartsheet
import openpyxl.styles
import openpyxl.worksheet.worksheet
import pkg_resources
import scipy.io
from PyQt5 import QtWidgets, QtCore, QtGui
from matplotlib.backend_bases import key_press_handler
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg, NavigationToolbar2QT
from matplotlib.figure import Figure
from sastool.io.credo_cct.header import Header
from scipy.misc import imread

from .headerpopup import HeaderPopup
from .mainwindow_ui import Ui_MainWindow
from .samplescalermodel import SampleScalerModel
from ..display import show_scattering_image, show_cmatrix, display_outlier_test_results, summarize_curves, \
    plot_vacuum_and_flux, make_transmission_table, make_exptimes_table
from ..export_table import export_table
from ..headermodel import HeaderModel
from ...core.processing.summarize import Summarizer
from ...core.utils.timeout import IdleFunction

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow, logging.Handler):
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self, None)
        logging.Handler.__init__(self, logger.level)
        self.rootdir = None
        self.processingprocess = None
        self.header_columns = HeaderModel.visiblecolumns
        self._lastsortcolumn = 1
        self._lastsortdirection = QtCore.Qt.AscendingOrder
        self.idlefcn = None
        logging.root.addHandler(self)
        self.setupUi(self)
        self.load_state()

    def emit(self, record: logging.LogRecord):
        msg = self.format(record)
        self.statusBar.showMessage(record.levelname + ': ' + msg)

    def load_state(self):
        statefile = os.path.join(appdirs.user_config_dir('cpt', 'CREDO', roaming=True), 'cpt.ini')
        config = configparser.ConfigParser()
        try:
            with open(statefile, 'rt', encoding='utf-8') as f:
                config.read_file(f)
        except FileNotFoundError:
            pass
        for widget, section, key in self._configwidgets:
            try:
                value = config[section][key]
            except KeyError:
                continue
            if isinstance(widget, QtWidgets.QLineEdit):
                widget.setText(value)
            elif isinstance(widget, QtWidgets.QComboBox):
                idx = widget.findText(config[section][key])
                if idx < 0:
                    continue
                widget.setCurrentIndex(idx)
            elif isinstance(widget, QtWidgets.QGroupBox):
                widget.setChecked(config[section][key].lower() == 'true')
            elif isinstance(widget, QtWidgets.QCheckBox):
                widget.setChecked(config[section][key].lower() == 'true')
            elif isinstance(widget, QtWidgets.QDoubleSpinBox):
                widget.setValue(float(value))
            elif isinstance(widget, QtWidgets.QSpinBox):
                widget.setValue(int(value))
            else:
                raise ValueError('Unknown widget type for section {} key {}: {}'.format(section, key, type(widget)))
        if self.rootDirLineEdit.text():
            self.setRootDir(self.rootDirLineEdit.text())
        if self.saveHDFLineEdit.text():
            self.processPushButton.setEnabled(True)
        try:
            self.header_columns = config['headerview']['fields'].split(';')
        except KeyError:
            pass
        try:
            self.setHDF5File(self.saveHDFLineEdit.text())
        except:
            pass

    def closeEvent(self, event: QtGui.QCloseEvent):
        self.save_state()
        event.accept()

    def save_state(self):
        configdir = appdirs.user_config_dir('cpt', 'CREDO', roaming=True)
        os.makedirs(configdir, exist_ok=True)
        statefile = os.path.join(configdir, 'cpt.ini')
        config = configparser.ConfigParser()
        for widget, section, item in self._configwidgets:
            if section not in config:
                config[section] = {}
            if isinstance(widget, QtWidgets.QLineEdit):
                config[section][item] = widget.text()
            elif isinstance(widget, (QtWidgets.QDoubleSpinBox, QtWidgets.QSpinBox)):
                config[section][item] = str(widget.value())
            elif isinstance(widget, (QtWidgets.QCheckBox, QtWidgets.QGroupBox)):
                config[section][item] = str(widget.isChecked())
            elif isinstance(widget, (QtWidgets.QComboBox,)):
                config[section][item] = widget.currentText()
            else:
                raise ValueError(
                    'Unknown widget type for config section {} item {}: {}'.format(section, item, type(widget)))
        config['headerview'] = {}
        config['headerview']['fields'] = ';'.join(self.header_columns)
        with open(statefile, 'wt', encoding='utf-8') as f:
            config.write(f)
        logger.info('Configuration saved to {}'.format(statefile))

    def onUiStyleChange(self):
        QtWidgets.QApplication.instance().setStyle(self.uiStyleComboBox.currentText())

    def setupUi(self, MainWindow: QtWidgets.QMainWindow):
        Ui_MainWindow.setupUi(self, MainWindow)
        self.uiStyleComboBox.addItems(QtWidgets.QStyleFactory.keys())
        currentkey = [k for k in QtWidgets.QStyleFactory.keys() if
                      k.upper() == QtWidgets.QApplication.instance().style().objectName().upper()][0]
        idx = self.uiStyleComboBox.findText(currentkey)
        self.uiStyleComboBox.setCurrentIndex(idx)
        self.uiStyleComboBox.currentIndexChanged.connect(self.onUiStyleChange)
        self.browseHDFPushButton.clicked.connect(self.onBrowseSaveFile)
        self.browsePushButton.clicked.connect(self.onBrowseRootDir)
        self.processPushButton.clicked.connect(self.onProcess)
        self.reloadPushButton.clicked.connect(self.onReload)
        self.headersTreeView.header().setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.headersTreeView.header().customContextMenuRequested.connect(self.onHeaderViewHeaderContextMenu)
        self.resultsDistanceSelectorComboBox.setEnabled(False)
        self.resultsSampleSelectorComboBox.setEnabled(False)
        self.resultsSampleSelectorComboBox.currentIndexChanged.connect(self.onResultsSampleSelected)
        self.resultsDistanceSelectorComboBox.currentIndexChanged.connect(self.onResultsDistanceSelected)
        self.plotCorMatPushButton.setEnabled(False)
        self.plotCorMatTestResultsPushButton.setEnabled(False)
        self.plotCurvesPushButton.setEnabled(False)
        self.plotImagePushButton.setEnabled(False)
        self.progressGroupBox.setVisible(False)
        self.plotCorMatTestResultsPushButton.clicked.connect(self.onPlotCorMatResults)
        self.plotCorMatPushButton.clicked.connect(self.onPlotCorMat)
        self.plotCurvesPushButton.clicked.connect(self.onPlotCurves)
        self.plotImagePushButton.clicked.connect(self.onPlotImage)
        self.qcVacuumFluxPushButton.clicked.connect(self.onQCVacuumFlux)
        self.qcTransmissionsPushButton.clicked.connect(self.onQCTransmissions)
        self.qcExposureTimesPushButton.clicked.connect(self.onQCExposureTimes)
        self.figure = Figure()
        self.canvas = FigureCanvasQTAgg(self.figure)
        vl = QtWidgets.QVBoxLayout(self.figureContainerWidget)
        vl.addWidget(self.canvas)
        self.figuretoolbar = NavigationToolbar2QT(self.canvas, self.figureContainerWidget)
        vl.addWidget(self.figuretoolbar)
        self._canvaskeypresshandler=self.canvas.mpl_connect('key_press_event', self.onCanvasKeyPressEvent)

        self.exportHeaderTablePushButton.clicked.connect(self.onExportHeadersTable)
        self.exportTablePushButton.clicked.connect(self.onExportTable)
        self.exportImageWidthUnitsComboBox.setCurrentIndex(self.exportImageWidthUnitsComboBox.findText('inch'))
        self.exportImageHeightUnitsComboBox.setCurrentIndex(self.exportImageHeightUnitsComboBox.findText('inch'))
        self.exportImageWidthDoubleSpinBox.setValue(matplotlib.rcParams['figure.figsize'][0])
        self.exportImageHeightDoubleSpinBox.setValue(matplotlib.rcParams['figure.figsize'][1])
        self.exportImageWidthUnitsComboBox.currentIndexChanged.connect(self.onImageUnitsChanged)
        self.exportImageHeightUnitsComboBox.currentIndexChanged.connect(self.onImageUnitsChanged)
        self.browseExportFolderPushButton.clicked.connect(self.onBrowseExportFolder)
        self.exportAveraged2DDataPushButton.clicked.connect(self.onExportAveraged2DData)
        self.exportAveraged2DGraphPushButton.clicked.connect(self.onExportAveraged2DGraph)
        self.exportAveragedCurvesDataPushButton.clicked.connect(self.onExportAveragedCurvesData)
        self.exportAveragedCurvesGraphPushButton.clicked.connect(self.onExportAveragedCurvesGraph)
        self.exportCorrelMatricesDataPushButton.clicked.connect(self.onExportCorrelMatricesData)
        self.exportCorrelMatricesGraphPushButton.clicked.connect(self.onExportCorrelMatricesGraph)
        self.headersTreeView.header().sectionClicked.connect(self.onHeaderTreeViewSortRequest)
        self.headersTreeView.setSortingEnabled(True)
        self.headersTreeView.sortByColumn(1, QtCore.Qt.AscendingOrder)
        self.ioProgressBar.setVisible(False)
        self.plot2DPushButton.clicked.connect(self.onPlot2D)
        self.plot1DPushButton.clicked.connect(self.onPlot1D)
        self.tablePlot1DPushButton.clicked.connect(self.onPlot1D)
        self.tablePlot2DPushButton.clicked.connect(self.onPlot2D)
        self.sampleNameListSelectAllPushButton.clicked.connect(self.onSelectAllSamples)
        self.sampleNameListDeselectAllPushButton.clicked.connect(self.onDeselectAllSamples)
        self.cmpSelectAllSamplesPushButton.clicked.connect(self.onCmpSelectAllSamples)
        self.cmpDeselectAllSamplesPushButton.clicked.connect(self.onCmpDeselectAllSamples)
        self.cmpPlotPushButton.clicked.connect(self.onCmpPlot)
        self._configwidgets = [
            (self.uiStyleComboBox, 'io', 'uistyle'),
            (self.saveHDFLineEdit, 'io', 'hdf5'),
            (self.rootDirLineEdit, 'io', 'datadir'),
            (self.exportFolderLineEdit, 'export', 'folder'),
            (self.cmpLabelLineEdit, 'curvecmp', 'legendformat'),
            (self.errorPropagationComboBox, 'processing', 'errorpropagation'),
            (self.abscissaErrorPropagationComboBox, 'processing', 'abscissaerrorpropagation'),
            (self.exportImageFormatComboBox, 'export', 'imageformat'),
            (self.exportImageHeightUnitsComboBox, 'export', 'imageheightunits'),
            (self.exportImageWidthUnitsComboBox, 'export', 'imagewidthunits'),
            (self.export1DDataFormatComboBox, 'export', 'onedimformat'),
            (self.cmpPlotTypeComboBox, 'curvecmp', 'plottype'),
            (self.firstFSNSpinBox, 'io', 'firstfsn'),
            (self.lastFSNSpinBox, 'io', 'lastfsn'),
            (self.stdMultiplierDoubleSpinBox, 'processing', 'std_multiplier'),
            (self.exportImageResolutionSpinBox, 'export', 'imagedpi'),
            (self.exportImageHeightDoubleSpinBox, 'export', 'imageheight'),
            (self.exportImageWidthDoubleSpinBox, 'export', 'imagewidth'),
            (self.qMaxDoubleSpinBox, 'processing', 'customqmax'),
            (self.qMinDoubleSpinBox, 'processing', 'customqmin'),
            (self.qBinCountSpinBox, 'processing', 'customqcount'),
            (self.logarithmicCorrelMatrixCheckBox, 'processing', 'logcorrelmatrix'),
            (self.sanitizeCurvesCheckBox, 'processing', 'sanitizecurves'),
            (self.logarithmicQCheckBox, 'processing', 'customqlogscale'),
            (self.qRangeOverrideGroupBox, 'processing', 'customq'),
            (self.plot1dShowMeanCurveCheckBox, 'persample', 'showmeancurve'),
            (self.plot1dShowBadCurvesCheckBox, 'persample', 'showbadcurves'),
            (self.plot1dShowGoodCurvesCheckBox, 'persample', 'showgoodcurves'),
            (self.plot1dLogarithmicXCheckBox, 'persample', 'logx'),
            (self.plot1dLogarithmicYCheckBox, 'persample', 'logy'),
            (self.plot2dShowMaskCheckBox, 'persample', 'showmask'),
            (self.plot2dShowCenterCheckBox, 'persample', 'showcenter'),
            (self.cmpErrorBarsCheckBox, 'curvecmp', 'errorbars'),
            (self.cmpLegendCheckBox, 'curvecmp', 'legend'),
        ]
        self.updateResults()

    def onCanvasKeyPressEvent(self, event):
        key_press_handler(event, self.canvas, self.figuretoolbar)

    def clearFigure(self):
        self.figure.clear()

    def onCmpPlot(self):
        try:
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as hdf5:
                self.clearFigure()
                ax = self.figure.add_subplot(1, 1, 1)
                arbunits = False
                for sn in self.cmpSampleModel.selectedSamples():
                    for dist in hdf5['Samples'][sn]:
                        grp = hdf5['Samples'][sn][dist]
                        curve = grp['curve']
                        factor = self.cmpSampleModel.factorForSample(sn)
                        if factor != 1:
                            arbunits = True
                        attrs = grp.attrs
                        try:
                            lbl = string.Formatter().vformat(self.cmpLabelLineEdit.text(), (), attrs)
                        except Exception as exc:
                            lbl = '--error-in-format-string--'
                        if self.cmpPlotTypeComboBox.currentText() == 'Kratky':
                            x = curve[:, 0]
                            y = curve[:, 1] * curve[:, 0] ** 2 * factor
                            dy = (4 * curve[:, 0] ** 2 * curve[:, 1] ** 2 * curve[:, 2] ** 2 +
                                  curve[:, 0] ** 4 * curve[:, 2] ** 2) ** 0.5
                            dx = curve[:, 3]
                        elif self.cmpPlotTypeComboBox.currentText() == 'Porod':
                            x = curve[:, 0]
                            y = curve[:, 1] * curve[:, 0] ** 4 * factor
                            dy = (16 * curve[:, 0] ** 6 * curve[:, 1] ** 2 * curve[:, 2] ** 2 +
                                  curve[:, 0] ** 8 * curve[:, 2] ** 2) ** 0.5
                            dx = curve[:, 3]
                        else:
                            x = curve[:, 0]
                            y = curve[:, 1] * factor
                            dy = curve[:, 2] * factor
                            dx = curve[:, 3]
                        if self.cmpErrorBarsCheckBox.isChecked():
                            ax.errorbar(x, y, dy, dx, label=lbl)
                        else:
                            ax.plot(x, y, label=lbl)
                ax.set_xlabel('q (nm$^{-1}$)')
                if arbunits:
                    if self.cmpPlotTypeComboBox.currentText() == 'Kratky':
                        ax.set_ylabel('q$^2 \\times$ relative intensity (nm$^{-2}\\times$arb. units)')
                    elif self.cmpPlotTypeComboBox.currentText() == 'Kratky':
                        ax.set_ylabel('q$^4 \\times$ relative intensity (nm$^{-4}\\times$arb. units)')
                    else:
                        ax.set_ylabel('Relative intensity (arb. units)')
                else:
                    if self.cmpPlotTypeComboBox.currentText() == 'Kratky':
                        ax.set_ylabel('$q^2\\times d\Sigma/d\Omega$ (nm$^{-2}$ cm$^{-1}$ sr$^{-1}$)')
                    elif self.cmpPlotTypeComboBox.currentText() == 'Porod':
                        ax.set_ylabel('$q^4\\times d\Sigma/d\Omega$ (nm$^{-4}$ cm$^{-1}$ sr$^{-1}$)')
                    else:
                        ax.set_ylabel('$d\Sigma/d\Omega$ (cm$^{-1}$ sr$^{-1}$)')
                if self.cmpPlotTypeComboBox.currentText() == 'log I vs. log q':
                    ax.set_xscale('log')
                    ax.set_yscale('log')
                elif self.cmpPlotTypeComboBox.currentText() == 'log I vs. q':
                    ax.set_xscale('linear')
                    ax.set_yscale('log')
                elif self.cmpPlotTypeComboBox.currentText() == 'I vs. log q':
                    ax.set_xscale('log')
                    ax.set_yscale('linear')
                elif self.cmpPlotTypeComboBox.currentText() == 'Guinier':
                    ax.set_xscale('power', exponent=2)
                    ax.set_yscale('log')
                elif self.cmpPlotTypeComboBox.currentText() == 'Kratky':
                    ax.set_xscale('linear')
                    ax.set_yscale('linear')
                elif self.cmpPlotTypeComboBox.currentText() == 'Porod':
                    ax.set_xscale('power', exponent=4)
                    ax.set_yscale('linear')
                ax.grid(True, which='both')
                if self.cmpLegendCheckBox.isChecked():
                    ax.legend(loc='best')
                self.figure.tight_layout()
                self.canvas.draw()
                self.tabWidget.setCurrentWidget(self.figureContainerWidget)
        except (FileNotFoundError, ValueError, OSError):
            return

    def onCmpSelectAllSamples(self):
        self.cmpSampleModel.selectAll()

    def onCmpDeselectAllSamples(self):
        self.cmpSampleModel.deselectAll()

    def onSelectAllSamples(self):
        self.sampleNameListWidget.selectAll()

    def onDeselectAllSamples(self):
        self.sampleNameListWidget.clearSelection()

    def onPlot2D(self):
        if self.sender() == self.plot2DPushButton:
            treeview = self.headersTreeView
        elif self.sender() == self.tablePlot2DPushButton:
            treeview = self.treeView
        else:
            return
        idx = treeview.currentIndex()
        if not idx.isValid():
            return
        try:
            fsn = treeview.model().getFSN(idx)
        except AttributeError:
            return
        summarizer = Summarizer([fsn], self.headermodel.eval2d_pathes, self.headermodel.eval2d_pathes,
                                self.headermodel.mask_pathes, self.saveHDFLineEdit.text(),
                                self.config['path']['prefixes']['crd'],
                                self.config['path']['fsndigits'])
        for x in summarizer.load_headers(logger=logger):
            pass
        ex = summarizer.load_exposure(fsn, logger=logger)
        self.clearFigure()
        ax = self.figure.add_subplot(1, 1, 1)
        ex.imshow(axes=ax, norm=matplotlib.colors.LogNorm())
        ax.set_xlabel('q (nm$^{-1}$)')
        ax.set_ylabel('q (nm$^{-1}$)')
        self.canvas.draw()
        self.tabWidget.setCurrentWidget(self.figureContainerWidget)

    def onPlot1D(self):
        if self.sender() == self.plot1DPushButton:
            treeview = self.headersTreeView
        elif self.sender() == self.tablePlot1DPushButton:
            treeview = self.treeView
        else:
            return
        idx = treeview.currentIndex()
        if not idx.isValid():
            return
        try:
            fsn = treeview.model().getFSN(idx)
        except AttributeError:
            return
        summarizer = Summarizer([fsn], self.headermodel.eval2d_pathes, self.headermodel.eval2d_pathes,
                                self.headermodel.mask_pathes, self.saveHDFLineEdit.text(),
                                self.config['path']['prefixes']['crd'],
                                self.config['path']['fsndigits'])
        for x in summarizer.load_headers():
            pass
        ex = summarizer.load_exposure(fsn)
        self.clearFigure()
        ax = self.figure.add_subplot(1, 1, 1)
        ex.radial_average().loglog(axes=ax)
        ax.set_xlabel('q (nm$^{-1}$)')
        ax.set_ylabel('Intensity (cm$^{-1}$ sr$^{-1}$)')
        self.canvas.draw()
        self.tabWidget.setCurrentWidget(self.figureContainerWidget)

    def onHeaderTreeViewSortRequest(self, section: int):
        self.headersTreeView.setSortingEnabled(True)
        if section == self._lastsortcolumn:
            if self._lastsortdirection == QtCore.Qt.AscendingOrder:
                self._lastsortdirection = QtCore.Qt.DescendingOrder
            else:
                self._lastsortdirection = QtCore.Qt.AscendingOrder
        else:
            self._lastsortcolumn = section
            self._lastsortdirection = QtCore.Qt.AscendingOrder
        self.headersTreeView.sortByColumn(self._lastsortcolumn, self._lastsortdirection)

    def onExportAveraged2DData(self):
        try:
            self.setEnabled(False)
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as hdf5:
                samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
                for sn in hdf5['Samples']:
                    if sn not in samplenames:
                        continue
                    for dist in hdf5['Samples'][sn]:
                        image = hdf5['Samples'][sn][dist]['image']
                        error = hdf5['Samples'][sn][dist]['image_uncertainty']
                        mask = hdf5['Samples'][sn][dist]['mask']
                        if self.export2DDataFormatComboBox.currentText() == 'Numpy':
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              '{}_{}.npz'.format(sn, dist.replace('.', '_')))
                            np.savez(fn, intensity=image, error=error, mask=mask)
                            logger.info('Wrote file {}'.format(fn))
                        elif self.export2DDataFormatComboBox.currentText() == 'Matlab':
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              '{}_{}.mat'.format(sn, dist.replace('.', '_')))
                            scipy.io.savemat(fn,
                                             {'intensity': image, 'error': error, 'mask': mask}, do_compression=True
                                             )
                            logger.info('Wrote file {}'.format(fn))
                        elif self.export2DDataFormatComboBox.currentText() == 'ASCII':
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              '{}_{}_intensity.txt'.format(sn, dist.replace('.', '_')))
                            np.savetxt(fn, image)
                            logger.info('Wrote file {}'.format(fn))
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              '{}_{}_error.txt'.format(sn, dist.replace('.', '_')))
                            np.savetxt(fn, error)
                            logger.info('Wrote file {}'.format(fn))
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              '{}_{}_mask.txt'.format(sn, dist.replace('.', '_')))
                            np.savetxt(fn, mask)
                            logger.info('Wrote file {}'.format(fn))
                        elif self.export2DDataFormatComboBox.currentText() == 'Gzip\'d ASCII':
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              '{}_{}_intensity.txt.gz'.format(sn, dist.replace('.', '_')))
                            np.savetxt(fn, image)
                            logger.info('Wrote file {}'.format(fn))
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              '{}_{}_error.txt.gz'.format(sn, dist.replace('.', '_')))
                            np.savetxt(fn, error)
                            logger.info('Wrote file {}'.format(fn))
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              '{}_{}_mask.txt.gz'.format(sn, dist.replace('.', '_')))
                            np.savetxt(fn, mask)
                            logger.info('Wrote file {}'.format(fn))
                        else:
                            raise ValueError(
                                'Unknown 2D file format: {}'.format(self.export2DDataFormatComboBox.currentText()))
        finally:
            self.setEnabled(True)

    def onExportAveraged2DGraph(self):
        width = self.exportImageWidthDoubleSpinBox.value()
        height = self.exportImageHeightDoubleSpinBox.value()
        if self.exportImageWidthUnitsComboBox.currentText() == 'cm':
            width /= 2.54
        if self.exportImageHeightUnitsComboBox.currentText() == 'cm':
            height /= 2.54

        fig = Figure(figsize=(width, height), dpi=self.exportImageResolutionSpinBox.value(), tight_layout=True)
        canvas = FigureCanvasQTAgg(fig)
        self.setEnabled(False)
        try:
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as hdf5:
                samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
                for sn in hdf5['Samples']:
                    if sn not in samplenames:
                        continue
                    for dist in hdf5['Samples'][sn]:
                        show_scattering_image(fig, hdf5['Samples'][sn][dist])
                        self.putlogo(fig)
                        fn = os.path.join(
                            self.exportFolderLineEdit.text(),
                            '{}_{}.{}'.format(sn, dist.replace('.', '_'), self.exportImageFormatComboBox.currentText())
                        )
                        fig.savefig(
                            fn,
                            dpi=self.exportImageResolutionSpinBox.value(),
                            format=self.exportImageFormatComboBox.currentText(),
                        )
                        logger.info('Wrote file {}'.format(fn))
        finally:
            self.setEnabled(True)
            del fig
            del canvas
            gc.collect()

    def onExportAveragedCurvesData(self):
        if self.export1DDataFormatComboBox.currentText() == 'ASCII (*.txt)':
            self.onExportAveragedCurvesDataASCII(with_qerror=True, extn='txt')
        elif self.export1DDataFormatComboBox.currentText() == 'ASCII (*.dat)':
            self.onExportAveragedCurvesDataASCII(with_qerror=False, extn='dat')
        elif self.export1DDataFormatComboBox.currentText().startswith('Excel 2007-'):
            self.onExportAveragedCurvesDataXLSX()

    def onExportAveragedCurvesDataXLSX(self):
        try:
            self.setEnabled(False)
            wb = openpyxl.Workbook()
            sheet = wb.get_active_sheet()
            chart = openpyxl.chart.ScatterChart()
            chart.y_axis.title = 'Intensity (1/cm * 1/sr)'
            chart.x_axis.title = 'q (1/nm)'
            chart.x_axis.scaling.logBase = 10
            chart.y_axis.scaling.logBase = 10
            chart.x_axis.majorTickMark = 'out'
            chart.y_axis.majorTickMark = 'out'
            chart.x_axis.minorTickMark = 'out'
            chart.y_axis.minorTickMark = 'out'
            assert isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet)
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as hdf5:
                samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
                i = 0
                qmins = []
                qmaxs = []
                Imins = []
                Imaxs = []
                for sn in sorted([s for s in hdf5['Samples'] if s in samplenames]):
                    for dist in hdf5['Samples'][sn]:
                        i += 1
                        data = hdf5['Samples'][sn][dist]['curve']
                        c = sheet.cell(row=1, column=(i - 1) * 4 + 1, value='{} @{} mm'.format(sn, dist))
                        c.font = openpyxl.styles.Font(bold=True)
                        c.alignment = openpyxl.styles.Alignment(horizontal='center')
                        sheet.merge_cells(start_row=1, end_row=1, start_column=(i - 1) * 4 + 1, end_column=i * 4)
                        sheet.cell(row=2, column=(i - 1) * 4 + 1, value='q (1/nm)').font = openpyxl.styles.Font(
                            bold=True)
                        sheet.cell(row=2, column=(i - 1) * 4 + 2, value='Intensity (1/cm)').font = openpyxl.styles.Font(
                            bold=True)
                        sheet.cell(row=2, column=(i - 1) * 4 + 3,
                                   value='Error of intensity (1/cm)').font = openpyxl.styles.Font(bold=True)
                        sheet.cell(row=2, column=(i - 1) * 4 + 4,
                                   value='Error of q (1/nm)').font = openpyxl.styles.Font(bold=True)
                        for j in range(data.shape[0]):
                            for k in range(data.shape[1]):
                                sheet.cell(row=3 + j, column=(i - 1) * 4 + 1 + k, value=data[j, k])
                        xvalues = openpyxl.chart.Reference(sheet, min_col=(i - 1) * 4 + 1, max_col=(i - 1) * 4 + 1,
                                                           min_row=3, max_row=data.shape[0] + 3)
                        yvalues = openpyxl.chart.Reference(sheet, min_col=(i - 1) * 4 + 2, max_col=(i - 1) * 4 + 2,
                                                           min_row=3, max_row=data.shape[0] + 3)
                        series = openpyxl.chart.Series(yvalues, xvalues, title='{} @{}'.format(sn, dist))
                        chart.series.append(series)
                        qmins.append(np.nanmin(data[:, 0]))
                        qmaxs.append(np.nanmax(data[:, 0]))
                        Imins.append(np.nanmin(data[:, 1]))
                        Imaxs.append(np.nanmax(data[:, 1]))

                chart.x_axis.scaling.min = min([x for x in qmins if x > 0])
                chart.x_axis.scaling.max = max(qmaxs)
                chart.y_axis.scaling.min = min([x for x in Imins if x > 0])
                chart.y_axis.scaling.max = max(Imaxs)
                # chart.style=13
                sheet.add_chart(chart)
                wb.save(os.path.join(self.exportFolderLineEdit.text(), 'SAXS_curves.xlsx'))
                logger.info('Wrote file {}'.format(os.path.join(self.exportFolderLineEdit.text(), 'SAXS_curves.xlsx')))
        finally:
            self.setEnabled(True)

    def onExportAveragedCurvesDataASCII(self, with_qerror=True, extn='txt'):
        try:
            self.setEnabled(False)
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as hdf5:
                samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
                for sn in hdf5['Samples']:
                    if sn not in samplenames:
                        continue
                    for dist in hdf5['Samples'][sn]:
                        data = hdf5['Samples'][sn][dist]['curve']
                        fn = os.path.join(
                            self.exportFolderLineEdit.text(),
                            '{}_{}.{}'.format(sn, dist.replace('.', '_'), extn))
                        with open(fn, 'wb') as f:
                            if with_qerror:
                                f.write('# q\tIntensity\tError\tqError\n'.encode('utf-8'))
                                np.savetxt(f, data)
                            else:
                                f.write('# q\tIntensity\tError\n'.encode('utf-8'))
                                np.savetxt(f, data[:,:3])
                        logger.info('Wrote file {}'.format(fn))
        finally:
            self.setEnabled(True)

    def onExportAveragedCurvesGraph(self):
        width = self.exportImageWidthDoubleSpinBox.value()
        height = self.exportImageHeightDoubleSpinBox.value()
        if self.exportImageWidthUnitsComboBox.currentText() == 'cm':
            width /= 2.54
        if self.exportImageHeightUnitsComboBox.currentText() == 'cm':
            height /= 2.54

        fig = Figure(figsize=(width, height), dpi=self.exportImageResolutionSpinBox.value(), tight_layout=True)
        canvas = FigureCanvasQTAgg(fig)
        axes = fig.add_subplot(1, 1, 1)
        try:
            self.setEnabled(False)
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as hdf5:
                samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
                for sn in hdf5['Samples']:
                    if sn not in samplenames:
                        continue
                    for dist in hdf5['Samples'][sn]:
                        data = hdf5['Samples'][sn][dist]['curve']
                        axes.loglog(data[:, 0], data[:, 1], label='{} @{} mm'.format(sn, dist.replace('.', '_')))
                axes.set_xlabel('$q$ (nm$^{-1}$')
                axes.set_ylabel('$d\Sigma/d\Omega$ (cm$^{-1}$sr$^{-1}$)')
                axes.legend(loc='best')
                axes.grid(True, which='both')
            self.putlogo(fig)
            fn = os.path.join(
                self.exportFolderLineEdit.text(),
                'SAXS_curves.{}'.format(self.exportImageFormatComboBox.currentText())
            )
            fig.savefig(
                fn,
                dpi=self.exportImageResolutionSpinBox.value(),
                format=self.exportImageFormatComboBox.currentText(),
            )
            logger.info('Wrote file {}'.format(fn))
        finally:
            self.setEnabled(True)
            del fig
            del canvas
            gc.collect()

    def onExportCorrelMatricesData(self):
        try:
            self.setEnabled(False)
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as hdf5:
                samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
                for sn in hdf5['Samples']:
                    if sn not in samplenames:
                        continue
                    for dist in hdf5['Samples'][sn]:
                        cmat = hdf5['Samples'][sn][dist]['correlmatrix']
                        if self.export2DDataFormatComboBox.currentText() == 'Numpy':
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              'correlmatrix_{}_{}.npz'.format(sn, dist.replace('.', '_')))
                            np.savez(fn, correlmatrix=cmat)
                            logger.info('Wrote file {}'.format(fn))
                        elif self.export2DDataFormatComboBox.currentText() == 'Matlab':
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              'correlmatrix_{}_{}.mat'.format(sn, dist.replace('.', '_')))
                            scipy.io.savemat(fn,
                                             {'correlmatrix': cmat}, do_compression=True
                                             )
                            logger.info('Wrote file {}'.format(fn))
                        elif self.export2DDataFormatComboBox.currentText() == 'ASCII':
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              'correlmatrix_{}_{}.txt'.format(sn, dist.replace('.', '_')))
                            np.savetxt(fn, cmat)
                            logger.info('Wrote file {}'.format(fn))
                        elif self.export2DDataFormatComboBox.currentText() == 'Gzip\'d ASCII':
                            fn = os.path.join(self.exportFolderLineEdit.text(),
                                              'correlmatrix_{}_{}.txt.gz'.format(sn, dist.replace('.', '_')))
                            np.savetxt(fn, cmat)
                            logger.info('Wrote file {}'.format(fn))
                        else:
                            raise ValueError(
                                'Unknown 2D file format: {}'.format(self.export2DDataFormatComboBox.currentText()))
        finally:
            self.setEnabled(True)

    def onExportCorrelMatricesGraph(self):
        width = self.exportImageWidthDoubleSpinBox.value()
        height = self.exportImageHeightDoubleSpinBox.value()
        if self.exportImageWidthUnitsComboBox.currentText() == 'cm':
            width /= 2.54
        if self.exportImageHeightUnitsComboBox.currentText() == 'cm':
            height /= 2.54

        fig = Figure(figsize=(width, height), dpi=self.exportImageResolutionSpinBox.value(), tight_layout=True)
        canvas = FigureCanvasQTAgg(fig)
        try:
            self.setEnabled(False)
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as hdf5:
                samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
                for sn in hdf5['Samples']:
                    if sn not in samplenames:
                        continue
                    for dist in hdf5['Samples'][sn]:
                        show_cmatrix(fig, hdf5['Samples'][sn][dist])
                        self.putlogo(fig)
                        fn = os.path.join(
                            self.exportFolderLineEdit.text(),
                            'correlmatrix_{}_{}.{}'.format(sn, dist, self.exportImageFormatComboBox.currentText())
                        )
                        fig.savefig(
                            fn,
                            dpi=self.exportImageResolutionSpinBox.value(),
                            format=self.exportImageFormatComboBox.currentText(),
                        )
                        logger.info('Wrote file {}'.format(fn))
        finally:
            self.setEnabled(True)
            del fig
            del canvas
            gc.collect()

    def onBrowseExportFolder(self):
        filename = QtWidgets.QFileDialog.getExistingDirectory(self, 'Select the folder to output files to...')
        if filename:
            self.exportFolderLineEdit.setText(os.path.normpath(filename))

    def onImageUnitsChanged(self):
        if self.sender() == self.exportImageWidthUnitsComboBox:
            spinbox = self.exportImageWidthDoubleSpinBox
        elif self.sender() == self.exportImageHeightUnitsComboBox:
            spinbox = self.exportImageHeightDoubleSpinBox
        else:
            return
        if self.sender().currentText() == 'inch':
            spinbox.setValue(spinbox.value() / 2.54)
        elif self.sender().currentText() == 'cm':
            spinbox.setValue(spinbox.value() * 2.54)

    def onQCExposureTimes(self):
        with h5py.File(self.saveHDFLineEdit.text(), 'r') as f:
            model = make_exptimes_table(f['Samples'])
        self.treeView.setModel(model)
        for c in range(model.columnCount()):
            self.treeView.resizeColumnToContents(c)
        self.tabWidget.setCurrentWidget(self.tableContainerWidget)

    def onQCTransmissions(self):
        with h5py.File(self.saveHDFLineEdit.text(), 'r') as f:
            model = make_transmission_table(f['Samples'])
        self.treeView.setModel(model)
        for c in range(model.columnCount()):
            self.treeView.resizeColumnToContents(c)
        self.tabWidget.setCurrentWidget(self.tableContainerWidget)

    def onQCVacuumFlux(self):
        with h5py.File(self.saveHDFLineEdit.text(), 'r') as f:
            plot_vacuum_and_flux(self.figure, f['Samples'], self.config['datareduction']['absintrefname'])
        self.putlogo()
        self.canvas.draw()
        self.tabWidget.setCurrentWidget(self.figureContainerWidget)

    def onExportHeadersTable(self):
        fname, fltr = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Write table to file...', '',
            'MS Excel 2007-2016 files (*.xlsx, *.xlsm);;MS Excel files (*.xls);;All files (*)',
            'MS Excel 2007-2016 files (*.xlsx, *.xlsm)')
        if not fname:
            return
        export_table(fname, self.headermodel)
        logger.info('Wrote file {}'.format(fname))

    def onExportTable(self):
        fname, fltr = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Write table to file...', '',
            'MS Excel 2007-2016 files (*.xlsx, *.xlsm);;MS Excel files (*.xls);;All files (*)',
            'MS Excel 2007-2016 files (*.xlsx, *.xlsm)')
        if not fname:
            return
        export_table(fname, self.treeView.model())
        logger.info('Wrote file {}'.format(fname))

    def onPlotCorMat(self):
        with getHDF5Group(self) as grp:
            show_cmatrix(self.figure, grp)
        self.putlogo()
        self.canvas.draw()
        self.tabWidget.setCurrentWidget(self.figureContainerWidget)

    def onPlotCorMatResults(self):
        with getHDF5Group(self) as grp:
            model = display_outlier_test_results(grp['curves'])
        self.treeView.setModel(model)
        for c in range(model.columnCount()):
            self.treeView.resizeColumnToContents(c)
        self.tabWidget.setCurrentWidget(self.tableContainerWidget)

    def onPlotCurves(self):
        with getHDF5Group(self) as grp:
            summarize_curves(self.figure, grp['curves'], self.plot1dShowGoodCurvesCheckBox.isChecked(),
                             self.plot1dShowBadCurvesCheckBox.isChecked(),
                             self.plot1dShowMeanCurveCheckBox.isChecked(),
                             self.plot1dLogarithmicXCheckBox.isChecked(),
                             self.plot1dLogarithmicYCheckBox.isChecked())
        self.putlogo()
        self.canvas.draw()
        self.tabWidget.setCurrentWidget(self.figureContainerWidget)

    def onPlotImage(self):
        with getHDF5Group(self) as grp:
            show_scattering_image(self.figure, grp, self.plot2dShowMaskCheckBox.isChecked(),
                                  self.plot2dShowCenterCheckBox.isChecked(), )
        self.putlogo()
        self.canvas.draw()
        self.tabWidget.setCurrentWidget(self.figureContainerWidget)

    def onResultsSampleSelected(self):
        if not len(self.resultsSampleSelectorComboBox):
            return
        try:
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as f:
                self.resultsDistanceSelectorComboBox.clear()
                self.resultsDistanceSelectorComboBox.addItems(
                    sorted(f['Samples'][self.resultsSampleSelectorComboBox.currentText()],
                           key=lambda x: float(x)))
                self.resultsDistanceSelectorComboBox.setCurrentIndex(0)
        except (FileNotFoundError, ValueError, OSError):
            return

    def onResultsDistanceSelected(self):
        self.plotImagePushButton.setEnabled(self.resultsDistanceSelectorComboBox.currentIndex() >= 0)
        self.plotCurvesPushButton.setEnabled(self.resultsDistanceSelectorComboBox.currentIndex() >= 0)
        self.plotCorMatPushButton.setEnabled(self.resultsDistanceSelectorComboBox.currentIndex() >= 0)
        self.plotCorMatTestResultsPushButton.setEnabled(self.resultsDistanceSelectorComboBox.currentIndex() >= 0)

    def resizeHeaderViewColumns(self):
        for i in range(self.headermodel.columnCount()):
            self.headersTreeView.resizeColumnToContents(i)

    def onHeaderViewHeaderContextMenu(self, position: QtCore.QPoint):
        self.headerpopup = HeaderPopup(
            self, self.headermodel.visiblecolumns,
            sorted([x[0] for x in inspect.getmembers(Header) if isinstance(x[1], property)]))
        self.headerpopup.applied.connect(self.onHeaderPopupApplied)
        self.headerpopup.show()
        self.headerpopup.move(self.mapToGlobal(position))
        self.headerpopup.closed.connect(self.onHeaderPopupDestroyed)
        self.resizeHeaderViewColumns()

    def onHeaderPopupDestroyed(self):
        del self.headerpopup

    def onHeaderPopupApplied(self):
        self.headermodel.visiblecolumns = self.headerpopup.fields
        self.header_columns = self.headerpopup.fields
        self.reloadHeaders()

    def reloadHeaders(self):
        self.headersTreeView.setSortingEnabled(False)
        self.headermodel.reloadHeaders()

    def onHeaderModelFSNLoaded(self, totalcount, currentcount, thisfsn):
        if totalcount == currentcount == thisfsn == 0:
            self.ioProgressBar.setVisible(False)
            self.setEnabled(True)
            self.headersTreeView.setSortingEnabled(True)
            self.headersTreeView.sortByColumn(self._lastsortcolumn, self._lastsortdirection)
            self.resizeHeaderViewColumns()
        else:
            if totalcount > 0 and not self.ioProgressBar.isVisible():
                self.setEnabled(False)
                self.ioProgressBar.setVisible(True)
                self.ioProgressBar.setMinimum(0)
                self.ioProgressBar.setMaximum(totalcount)
                self.ioProgressBar.setFormat('Loading headers...')
            self.ioProgressBar.setValue(currentcount)

    def onReload(self):
        try:
            self.setEnabled(False)
            self.statusBar.showMessage('Loading headers, please wait...')
            newheadermodel = HeaderModel(
                self,
                self.rootdir,
                self.config['path']['prefixes']['crd'],
                self.firstFSNSpinBox.value(),
                self.lastFSNSpinBox.value(),
                self.header_columns,
                os.path.join(appdirs.user_state_dir('cpt', 'CREDO', roaming=True), 'badfsns')
            )
            newheadermodel.fsnloaded.connect(self.onHeaderModelFSNLoaded)
            newheadermodel.reloadHeaders()
            self.headersTreeView.setSortingEnabled(False)
            self.headersTreeView.setModel(newheadermodel)
            if hasattr(self, 'headermodel'):
                self.headermodel.fsnloaded.disconnect(self.onHeaderModelFSNLoaded)
                self.headermodel.cleanup()
                del self.headermodel
            self.headermodel = newheadermodel
            self.resizeHeaderViewColumns()
            # self.statusBar.showMessage('Headers loaded.')
            self.headersTreeView.setSortingEnabled(True)
            self.headersTreeView.sortByColumn(self._lastsortcolumn, self._lastsortdirection)
        finally:
            self.setEnabled(True)

    def onProcess(self):
        self.queue = queue.Queue()
        badfsns = self.headermodel.get_badfsns()
        if self.qRangeOverrideGroupBox.isChecked():
            qmin = self.qMinDoubleSpinBox.value()
            qmax = self.qMaxDoubleSpinBox.value()
            nq = self.qBinCountSpinBox.value()
            if self.logarithmicQCheckBox.isChecked():
                qrange = np.logspace(np.log10(qmin), np.log10(qmax), nq)
            else:
                qrange = np.linspace(qmin, qmax, nq)
        else:
            qrange = None
        kwargs = {
            'fsns': [x for x in range(self.firstFSNSpinBox.value(), self.lastFSNSpinBox.value() + 1) if
                     x not in badfsns],
            'exppath': self.headermodel.eval2d_pathes,
            'parampath': self.headermodel.eval2d_pathes,
            'maskpath': self.headermodel.mask_pathes,
            'outputfile': self.saveHDFLineEdit.text(),
            'prefix': self.config['path']['prefixes']['crd'],
            'ndigits': self.config['path']['fsndigits'],
            'errorpropagation': self.errorPropagationComboBox.currentIndex(),
            'abscissaerrorpropagation': self.abscissaErrorPropagationComboBox.currentIndex(),
            'sanitize_curves': self.sanitizeCurvesCheckBox.isChecked(),
            'logarithmic_correlmatrix': self.logarithmicCorrelMatrixCheckBox.isChecked(),
            'std_multiplier': self.stdMultiplierDoubleSpinBox.value(),
            'queue': self.queue,
            'qrange': qrange,
        }
        self.processingprocess = threading.Thread(target=self.do_processing, name='Summarization', kwargs=kwargs)
        self.idlefcn = IdleFunction(self.check_processing_progress, 100)
        self.processingprocess.start()
        self.progressGroupBox.setVisible(True)
        self.processPushButton.setEnabled(False)
        self.toolBox.setEnabled(False)

    def do_processing(self, queue: queue.Queue,
                      fsns: Iterable[int], exppath: Iterable[str], parampath: Iterable[str], maskpath: Iterable[str],
                      outputfile: str, prefix: str, ndigits: int, errorpropagation: int, abscissaerrorpropagation: int,
                      sanitize_curves: bool, logarithmic_correlmatrix: bool, std_multiplier: int,
                      qrange: Optional[np.ndarray]):
        s = Summarizer(fsns, exppath, parampath, maskpath, outputfile, prefix, ndigits, errorpropagation,
                       abscissaerrorpropagation, sanitize_curves, logarithmic_correlmatrix, std_multiplier, qrange)
        queue.put_nowait(('__init_loadheaders__', len(fsns)))
        for msg1, msg2 in s.load_headers(yield_messages=True):
            queue.put_nowait((msg1, msg2))
        for msg1, msg2 in s.summarize(True, yield_messages=True):
            queue.put_nowait((msg1, msg2))
        queue.put_nowait(('__done__', None))
        return True

    def check_processing_progress(self):
        if self.processingprocess is None:
            self.idlefcn = None
            self.queue = None
            self.processPushButton.setEnabled(True)
            self.toolBox.setEnabled(True)
            return False
        try:
            assert isinstance(self.queue, queue.Queue)
            for imessage in range(20):
                msg1, msg2 = self.queue.get_nowait()
                if msg1 == '__init_summarize__':
                    assert isinstance(msg2, int)  # msg2 is the number of samples times the number of distances
                    self.progressBar1.setMinimum(0)
                    self.progressBar1.setMaximum(msg2)
                    self.progressbar1StatusLabel.setText('')
                    self.progressbar1TitleLabel.setText('Processed sample:')
                    self.progressbar2TitleLabel.setText('Processed exposure:')
                    self.progressBar2.setMinimum(0)
                    self.progressBar2.setMaximum(0)
                    self.progressBar2.setVisible(True)
                    self.progressbar2TitleLabel.setVisible(True)
                    self.progressbar2StatusLabel.setVisible(True)
                elif msg1 == '__init_loadheaders__':
                    assert isinstance(msg2, int)  # the number of headers to load
                    self.progressBar2.setVisible(False)
                    self.progressbar2StatusLabel.setVisible(False)
                    self.progressbar2TitleLabel.setVisible(False)
                    self.progressbar1TitleLabel.setText('Loaded header:')
                    self.progressbar1StatusLabel.setText('')
                    self.progressBar1.setMinimum(0)
                    self.progressBar1.setMaximum(msg2)
                    self.progressBar1.setValue(0)
                elif msg1 in ['__header_loaded__', '__header_notfound__']:
                    # loaded a header for FSN
                    assert isinstance(msg2, int)  # msg2 is the FSN of the header just loaded or not found.
                    self.progressBar1.setValue(self.progressBar1.value() + 1)
                    self.progressbar1StatusLabel.setText('{:d}'.format(msg2))
                elif msg1 in ['__exposure_loaded__', '__exposure_notfound__']:
                    self.progressBar2.setValue(self.progressBar2.value() + 1)
                    self.progressbar2StatusLabel.setText('{:d}'.format(msg2))
                elif msg1 == '__init_collect_exposures__':
                    self.progressbar2TitleLabel.setText('Processed exposure:')
                    self.progressbar2StatusLabel.setText('')
                    self.progressBar2.setMinimum(0)
                    self.progressBar2.setMaximum(msg2)
                elif msg1 == '__init_stabilityassessment__':
                    self.progressBar2.setMinimum(0)
                    self.progressBar2.setMaximum(0)
                    self.progressbar2StatusLabel.setText('')
                    self.progressbar2TitleLabel.setText('Stability assessment...')
                elif msg1 == '__done__':
                    self.processingprocess.join()
                    self.processingprocess = None
                    self.progressGroupBox.setVisible(False)
                    self.idlefcn.stop()
                    self.idlefcn = None
                    self.processPushButton.setEnabled(True)
                    self.toolBox.setEnabled(True)
                    self.updateResults(processingfinished=True)
                    return False
                elif msg1.startswith('__'):
                    pass
                else:
                    assert isinstance(msg1, str)  # sample
                    assert isinstance(msg2, float)  # distance
                    self.progressBar1.setValue(self.progressBar1.value() + 1)
                    self.progressbar1StatusLabel.setText('{} ({:.2f} mm)'.format(msg1, msg2))
        except queue.Empty:
            return True
        return True

    def setRootDir(self, rootdir: str):
        self.rootdir = rootdir
        configfile = os.path.join(self.rootdir, 'config', 'cct.pickle')
        try:
            with open(configfile, 'rb') as f:
                self.config = pickle.load(f)
        except FileNotFoundError:
            QtWidgets.QMessageBox.critical(self, 'Error while loading config file',
                                           'Error while loading config file: {} not found.'.format(configfile))
            return False
        except pickle.PickleError:
            QtWidgets.QMessageBox.critical(self, 'Error while loading config file',
                                           'Error while loading config file: {} is malformed'.format(configfile))
            return False
        self.firstFSNSpinBox.setEnabled(True)
        self.lastFSNSpinBox.setEnabled(True)
        self.reloadPushButton.setEnabled(True)
        self.rootDirLineEdit.setText(os.path.normpath(rootdir))
        return True

    def onBrowseRootDir(self):
        filename = QtWidgets.QFileDialog.getExistingDirectory(self, 'Open CREDO data directory')
        if filename:
            self.setRootDir(filename)

    def setHDF5File(self, filename):
        self.saveHDFLineEdit.setText(os.path.normpath(filename))
        self.processPushButton.setEnabled(True)
        self.updateResults()

    def onBrowseSaveFile(self):
        filename, fltr = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save processed results to...', '',
            'HDF5 files (*.h5 *.hdf5);;All files (*)',
            'HDF5 files (*.h5 *.hdf5)')
        if filename is None:
            return
        if not filename.endswith('.h5') and not filename.endswith('.hdf5'):
            filename = filename + '.h5'
        self.setHDF5File(filename)

    def updateResults(self, processingfinished=False):
        try:
            with h5py.File(self.saveHDFLineEdit.text(), 'r') as f:
                samples = sorted(f['Samples'].keys())
                if self.autoMarkBadExposuresCheckBox.isChecked():
                    newbadfsns = []
                    for sn in f['Samples']:
                        for dist in f['Samples'][sn]:
                            for dset in f['Samples'][sn][dist]['curves'].values():
                                if dset.attrs['correlmat_bad']:
                                    newbadfsns.append(dset.attrs['fsn'])
                    try:
                        self.headermodel.update_badfsns(newbadfsns)
                    except AttributeError:
                        pass
                    if processingfinished:
                        if newbadfsns:
                            QtWidgets.QMessageBox.information(
                                self, 'Processing finished',
                                'Found and marked new bad exposures:\n' + ', '.join(
                                    [str(f) for f in sorted(newbadfsns)]),
                            )
                        else:
                            QtWidgets.QMessageBox.information(
                                self, 'Processing finished',
                                'No new bad exposures found',
                            )
        except (FileNotFoundError, ValueError, OSError):
            return
        self.resultsSampleSelectorComboBox.clear()
        self.resultsSampleSelectorComboBox.addItems(samples)
        self.resultsSampleSelectorComboBox.setCurrentIndex(0)
        self.resultsSampleSelectorComboBox.setEnabled(True)
        self.resultsDistanceSelectorComboBox.setEnabled(True)
        self.sampleNameListWidget.clear()
        self.sampleNameListWidget.addItems(samples)
        self.sampleNameListWidget.selectAll()
        self.cmpSampleModel = SampleScalerModel(samples)
        self.cmpSampleTreeView.setModel(self.cmpSampleModel)
        for i in range(self.cmpSampleModel.columnCount()):
            self.cmpSampleTreeView.resizeColumnToContents(i)

    def putlogo(self, figure: Optional[Figure] = None):
        if figure is None:
            figure = self.figure
        if not hasattr(self, '_logodata'):
            self._logodata = imread(pkg_resources.resource_filename('cct', 'resource/credo_logo.png'), flatten=True)[
                             ::4, ::4]
        figure.figimage(self._logodata, 10, 10, cmap='gray', zorder=-10)


class getHDF5Group:
    def __init__(self, mainwin: MainWindow, sample: Optional[str] = None, distance: Union[str, float, None] = None):
        self.mainwin = mainwin
        if sample is None:
            sample = self.mainwin.resultsSampleSelectorComboBox.currentText()
        if distance is None:
            distance = self.mainwin.resultsDistanceSelectorComboBox.currentText()
        if isinstance(distance, float):
            distance = '{:.2f}'.format(distance)
        self.sample = sample
        self.distance = distance
        self.hdf5 = None

    def __enter__(self):
        self.hdf5 = h5py.File(self.mainwin.saveHDFLineEdit.text(), 'r+')
        return self.hdf5['Samples'][self.sample][self.distance]

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.hdf5.close()
        self.hdf5 = None
