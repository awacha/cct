import gc
import logging
import os

import matplotlib
import numpy as np
import openpyxl
import openpyxl.chart
import openpyxl.styles
import openpyxl.worksheet.worksheet
import pkg_resources
import scipy.io
from imageio import imread
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.figure import Figure

from .exporttool_ui import Ui_Form
from ..toolbase import ToolBase
from ...display import show_cmatrix, show_scattering_image
from ...export import export2D

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


class ExportTool(ToolBase, Ui_Form):
    def setupUi(self, Form):
        super().setupUi(Form)
        self.exportImageWidthUnitsComboBox.setCurrentIndex(self.exportImageWidthUnitsComboBox.findText('inch'))
        self.exportImageHeightUnitsComboBox.setCurrentIndex(self.exportImageHeightUnitsComboBox.findText('inch'))
        self.exportImageWidthDoubleSpinBox.setValue(matplotlib.rcParams['figure.figsize'][0])
        self.exportImageHeightDoubleSpinBox.setValue(matplotlib.rcParams['figure.figsize'][1])
        self.exportImageWidthUnitsComboBox.currentIndexChanged.connect(self.onImageUnitsChanged)
        self.exportImageHeightUnitsComboBox.currentIndexChanged.connect(self.onImageUnitsChanged)
        self.exportAveraged2DDataPushButton.clicked.connect(self.onExportAveraged2DData)
        self.exportAveraged2DGraphPushButton.clicked.connect(self.onExportAveraged2DGraph)
        self.exportAveragedCurvesDataPushButton.clicked.connect(self.onExportAveragedCurvesData)
        self.exportAveragedCurvesGraphPushButton.clicked.connect(self.onExportAveragedCurvesGraph)
        self.exportCorrelMatricesDataPushButton.clicked.connect(self.onExportCorrelMatricesData)
        self.exportCorrelMatricesGraphPushButton.clicked.connect(self.onExportCorrelMatricesGraph)
        self.sampleNameListSelectAllPushButton.clicked.connect(self.onSelectAllSamples)
        self.sampleNameListDeselectAllPushButton.clicked.connect(self.onDeselectAllSamples)
        self.configWidgets = [
            (self.exportImageFormatComboBox, 'export', 'imageformat'),
            (self.exportImageHeightUnitsComboBox, 'export', 'imageheightunits'),
            (self.exportImageWidthUnitsComboBox, 'export', 'imagewidthunits'),
            (self.export1DDataFormatComboBox, 'export', 'onedimformat'),
            (self.exportImageResolutionSpinBox, 'export', 'imagedpi'),
            (self.exportImageHeightDoubleSpinBox, 'export', 'imageheight'),
            (self.exportImageWidthDoubleSpinBox, 'export', 'imagewidth'),
        ]

    def onSelectAllSamples(self):
        self.sampleNameListWidget.selectAll()

    def onDeselectAllSamples(self):
        self.sampleNameListWidget.clearSelection()

    def onExportAveraged2DData(self):
        try:
            self.busy.emit(True)
            samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
            for sn in self.h5GetSamples():
                if sn not in samplenames:
                    continue
                for dist in self.h5GetDistances(sn):
                    with self.getHDF5Group(sn, dist) as grp:
                        basename = os.path.join(self.exportFolder, '{}_{}'.format(sn, dist.replace('.', '_')))
                        if self.export2DDataFormatComboBox.currentText() == 'Numpy':
                            filelist = export2D.exportNumpy(basename, grp)
                        elif self.export2DDataFormatComboBox.currentText() == 'Matlab':
                            filelist = export2D.exportMatlab(basename, grp)
                        elif self.export2DDataFormatComboBox.currentText() == 'ASCII':
                            filelist = export2D.exportAscii(basename, grp, gzip=False)
                        elif self.export2DDataFormatComboBox.currentText() == 'Gzip\'d ASCII':
                            filelist = export2D.exportAscii(basename, grp, gzip=True)
                        else:
                            raise ValueError(
                                'Unknown 2D file format: {}'.format(self.export2DDataFormatComboBox.currentText()))
                    logger.info('Wrote file(s): {}'.format(', '.join([f for f in filelist])))
        finally:
            self.busy.emit(False)

    def onExportAveraged2DGraph(self):
        width = self.exportImageWidthDoubleSpinBox.value()
        height = self.exportImageHeightDoubleSpinBox.value()
        if self.exportImageWidthUnitsComboBox.currentText() == 'cm':
            width /= 2.54
        if self.exportImageHeightUnitsComboBox.currentText() == 'cm':
            height /= 2.54

        fig = Figure(figsize=(width, height), dpi=self.exportImageResolutionSpinBox.value(), tight_layout=True)
        canvas = FigureCanvasQTAgg(fig)
        self.busy.emit(True)
        try:
            samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
            for sn in self.h5GetSamples():
                if sn not in samplenames:
                    continue
                for dist in self.h5GetDistances(sn):
                    with self.getHDF5Group(sn, dist) as grp:
                        show_scattering_image(fig, grp)
                    self.putlogo(fig)
                    fn = os.path.join(
                        self.exportFolder,
                        '{}_{}.{}'.format(sn, dist.replace('.', '_'), self.exportImageFormatComboBox.currentText())
                    )
                    fig.savefig(
                        fn,
                        dpi=self.exportImageResolutionSpinBox.value(),
                        format=self.exportImageFormatComboBox.currentText(),
                    )
                    logger.info('Wrote file {}'.format(fn))
        finally:
            self.busy.emit(False)
            del fig
            del canvas
            gc.collect()

    def onExportAveragedCurvesData(self):
        if self.export1DDataFormatComboBox.currentText() == 'ASCII (*.txt)':
            self.onExportAveragedCurvesDataASCII(with_qerror=True, extn='txt')
        elif self.export1DDataFormatComboBox.currentText() == 'ASCII (*.dat)':
            self.onExportAveragedCurvesDataASCII(with_qerror=False, extn='dat')
        elif self.export1DDataFormatComboBox.currentText().startswith('Excel 2007-'):
            self.onExportAveragedCurvesDataXLSX()
        elif self.export1DDataFormatComboBox.currentText().startswith('RSR (*.rsr)'):
            self.onExportAveragedCurvesDataRSR()

    def onExportAveragedCurvesDataRSR(self):
        try:
            self.busy.emit(True)
            samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
            for sn in self.h5GetSamples():
                if sn not in samplenames:
                    continue
                for dist in self.h5GetDistances(sn):
                    with self.getHDF5Group(sn, dist) as grp:
                        data = np.array(grp['curve'])
                    fn = os.path.join(
                        self.exportFolder,
                        '{}_{}.rsr'.format(sn, dist.replace('.', '_')))
                    with open(fn, 'wt', newline='\r\n') as f:
                        f.write(' TIME\n 1.0\n {:d}\n'.format(data.shape[0]))
                        for i in range(data.shape[0]):
                            f.write(' {:.9f} {:.9f} {:.9f} 1\n'.format(data[i,0]/10., data[i,1], data[i,2]))
                    logger.info('Wrote file {}'.format(fn))
        finally:
            self.busy.emit(False)

    def onExportAveragedCurvesDataXLSX(self):
        try:
            self.busy.emit(True)
            wb = openpyxl.Workbook()
            sheet = wb.get_active_sheet()
            chart = openpyxl.chart.ScatterChart()
            chart.y_axis.title = 'Intensity (1/cm * 1/sr)'
            chart.x_axis.title = 'q (1/nm)'
            chart.x_axis.scaling.logBase = 10
            chart.y_axis.scaling.logBase = 10
            chart.x_axis.majorTickMark = 'out'
            chart.y_axis.majorTickMark = 'out'
            chart.x_axis.minorTickMark = 'out'
            chart.y_axis.minorTickMark = 'out'
            assert isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet)
            samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
            i = 0
            qmins = []
            qmaxs = []
            Imins = []
            Imaxs = []
            for sn in sorted([s for s in self.h5GetSamples() if s in samplenames]):
                for dist in self.h5GetDistances(sn):
                    i += 1
                    with self.getHDF5Group(sn, dist) as grp:
                        data = np.array(grp['curve'])
                    c = sheet.cell(row=1, column=(i - 1) * 4 + 1, value='{} @{} mm'.format(sn, dist))
                    c.font = openpyxl.styles.Font(bold=True)
                    c.alignment = openpyxl.styles.Alignment(horizontal='center')
                    sheet.merge_cells(start_row=1, end_row=1, start_column=(i - 1) * 4 + 1, end_column=i * 4)
                    sheet.cell(row=2, column=(i - 1) * 4 + 1, value='q (1/nm)').font = openpyxl.styles.Font(
                        bold=True)
                    sheet.cell(row=2, column=(i - 1) * 4 + 2, value='Intensity (1/cm)').font = openpyxl.styles.Font(
                        bold=True)
                    sheet.cell(row=2, column=(i - 1) * 4 + 3,
                               value='Error of intensity (1/cm)').font = openpyxl.styles.Font(bold=True)
                    sheet.cell(row=2, column=(i - 1) * 4 + 4,
                               value='Error of q (1/nm)').font = openpyxl.styles.Font(bold=True)
                    for j in range(data.shape[0]):
                        for k in range(data.shape[1]):
                            sheet.cell(row=3 + j, column=(i - 1) * 4 + 1 + k, value=data[j, k])
                    xvalues = openpyxl.chart.Reference(sheet, min_col=(i - 1) * 4 + 1, max_col=(i - 1) * 4 + 1,
                                                       min_row=3, max_row=data.shape[0] + 3)
                    yvalues = openpyxl.chart.Reference(sheet, min_col=(i - 1) * 4 + 2, max_col=(i - 1) * 4 + 2,
                                                       min_row=3, max_row=data.shape[0] + 3)
                    series = openpyxl.chart.Series(yvalues, xvalues, title='{} @{}'.format(sn, dist))
                    chart.series.append(series)
                    qmins.append(np.nanmin(data[:, 0]))
                    qmaxs.append(np.nanmax(data[:, 0]))
                    Imins.append(np.nanmin(data[:, 1]))
                    Imaxs.append(np.nanmax(data[:, 1]))

            chart.x_axis.scaling.min = min([x for x in qmins if x > 0])
            chart.x_axis.scaling.max = max(qmaxs)
            chart.y_axis.scaling.min = min([x for x in Imins if x > 0])
            chart.y_axis.scaling.max = max(Imaxs)
            # chart.style=13
            sheet.add_chart(chart)
            wb.save(os.path.join(self.exportFolder, 'SAXS_curves.xlsx'))
            logger.info('Wrote file {}'.format(os.path.join(self.exportFolder, 'SAXS_curves.xlsx')))
        finally:
            self.busy.emit(False)

    def onExportAveragedCurvesDataASCII(self, with_qerror=True, extn='txt'):
        try:
            self.busy.emit(True)
            samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
            for sn in self.h5GetSamples():
                if sn not in samplenames:
                    continue
                for dist in self.h5GetDistances(sn):
                    with self.getHDF5Group(sn, dist) as grp:
                        data = np.array(grp['curve'])
                    fn = os.path.join(
                        self.exportFolder,
                        '{}_{}.{}'.format(sn, dist.replace('.', '_'), extn))
                    with open(fn, 'wb') as f:
                        if with_qerror:
                            f.write('# q\tIntensity\tError\tqError\n'.encode('utf-8'))
                            np.savetxt(f, data)
                        else:
                            f.write('# q\tIntensity\tError\n'.encode('utf-8'))
                            np.savetxt(f, data[:, :3])
                    logger.info('Wrote file {}'.format(fn))
        finally:
            self.busy.emit(False)

    def onExportAveragedCurvesGraph(self):
        width = self.exportImageWidthDoubleSpinBox.value()
        height = self.exportImageHeightDoubleSpinBox.value()
        if self.exportImageWidthUnitsComboBox.currentText() == 'cm':
            width /= 2.54
        if self.exportImageHeightUnitsComboBox.currentText() == 'cm':
            height /= 2.54

        fig = Figure(figsize=(width, height), dpi=self.exportImageResolutionSpinBox.value(), tight_layout=True)
        canvas = FigureCanvasQTAgg(fig)
        axes = fig.add_subplot(1, 1, 1)
        try:
            self.busy.emit(True)
            samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
            for sn in self.h5GetSamples():
                if sn not in samplenames:
                    continue
                for dist in self.h5GetDistances(sn):
                    with self.getHDF5Group(sn, dist) as grp:
                        data = np.array(grp['curve'])
                    axes.loglog(data[:, 0], data[:, 1], label='{} @{} mm'.format(sn, dist.replace('.', '_')))
            axes.set_xlabel('$q$ (nm$^{-1}$')
            axes.set_ylabel('$d\Sigma/d\Omega$ (cm$^{-1}$sr$^{-1}$)')
            axes.legend(loc='best')
            axes.grid(True, which='both')
            self.putlogo(fig)
            fn = os.path.join(
                self.exportFolder,
                'SAXS_curves.{}'.format(self.exportImageFormatComboBox.currentText())
            )
            fig.savefig(
                fn,
                dpi=self.exportImageResolutionSpinBox.value(),
                format=self.exportImageFormatComboBox.currentText(),
            )
            logger.info('Wrote file {}'.format(fn))
        finally:
            self.busy.emit(False)
            del fig
            del canvas
            gc.collect()

    def onExportCorrelMatricesData(self):
        try:
            self.busy.emit(True)
            samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
            for sn in self.h5GetSamples():
                if sn not in samplenames:
                    continue
                for dist in self.h5GetDistances(sn):
                    try:
                        with self.getHDF5Group(sn, dist) as grp:
                            cmat = np.array(grp['correlmatrix'])
                    except KeyError:
                        logger.warning(
                            'Cannot export correlation matrix for sample {} at {} mm: no correlation matrix. Possibly a subtracted measurement.'.format(
                                sn, dist))
                        continue
                    if self.export2DDataFormatComboBox.currentText() == 'Numpy':
                        fn = os.path.join(self.exportFolder,
                                          'correlmatrix_{}_{}.npz'.format(sn, dist.replace('.', '_')))
                        np.savez(fn, correlmatrix=cmat)
                        logger.info('Wrote file {}'.format(fn))
                    elif self.export2DDataFormatComboBox.currentText() == 'Matlab':
                        fn = os.path.join(self.exportFolder,
                                          'correlmatrix_{}_{}.mat'.format(sn, dist.replace('.', '_')))
                        scipy.io.savemat(fn,
                                         {'correlmatrix': cmat}, do_compression=True
                                         )
                        logger.info('Wrote file {}'.format(fn))
                    elif self.export2DDataFormatComboBox.currentText() == 'ASCII':
                        fn = os.path.join(self.exportFolder,
                                          'correlmatrix_{}_{}.txt'.format(sn, dist.replace('.', '_')))
                        np.savetxt(fn, cmat)
                        logger.info('Wrote file {}'.format(fn))
                    elif self.export2DDataFormatComboBox.currentText() == 'Gzip\'d ASCII':
                        fn = os.path.join(self.exportFolder,
                                          'correlmatrix_{}_{}.txt.gz'.format(sn, dist.replace('.', '_')))
                        np.savetxt(fn, cmat)
                        logger.info('Wrote file {}'.format(fn))
                    else:
                        raise ValueError(
                            'Unknown 2D file format: {}'.format(self.export2DDataFormatComboBox.currentText()))
        finally:
            self.busy.emit(False)

    def onExportCorrelMatricesGraph(self):
        width = self.exportImageWidthDoubleSpinBox.value()
        height = self.exportImageHeightDoubleSpinBox.value()
        if self.exportImageWidthUnitsComboBox.currentText() == 'cm':
            width /= 2.54
        if self.exportImageHeightUnitsComboBox.currentText() == 'cm':
            height /= 2.54

        fig = Figure(figsize=(width, height), dpi=self.exportImageResolutionSpinBox.value(), tight_layout=True)
        canvas = FigureCanvasQTAgg(fig)
        try:
            self.busy.emit(True)
            samplenames = [item.text() for item in self.sampleNameListWidget.selectedItems()]
            for sn in self.h5GetSamples():
                if sn not in samplenames:
                    continue
                for dist in self.h5GetDistances(sn):
                    with self.getHDF5Group(sn, dist) as grp:
                        if 'correlmatrix' not in grp:
                            logger.warning('No correlation matrix for sample {} at {} mm.'.format(sn, dist))
                            continue
                        show_cmatrix(fig, grp)
                    self.putlogo(fig)
                    fn = os.path.join(
                        self.exportFolder,
                        'correlmatrix_{}_{}.{}'.format(sn, dist, self.exportImageFormatComboBox.currentText())
                    )
                    fig.savefig(
                        fn,
                        dpi=self.exportImageResolutionSpinBox.value(),
                        format=self.exportImageFormatComboBox.currentText(),
                    )
                    logger.info('Wrote file {}'.format(fn))
        finally:
            self.busy.emit(False)
            del fig
            del canvas
            gc.collect()

    def onImageUnitsChanged(self):
        if self.sender() == self.exportImageWidthUnitsComboBox:
            spinbox = self.exportImageWidthDoubleSpinBox
        elif self.sender() == self.exportImageHeightUnitsComboBox:
            spinbox = self.exportImageHeightDoubleSpinBox
        else:
            return
        if self.sender().currentText() == 'inch':
            spinbox.setValue(spinbox.value() / 2.54)
        elif self.sender().currentText() == 'cm':
            spinbox.setValue(spinbox.value() * 2.54)

    def setH5FileName(self, h5filename: str):
        super().setH5FileName(h5filename)
        self.sampleNameListWidget.clear()
        self.sampleNameListWidget.addItems(self.h5GetSamples())
        self.sampleNameListWidget.selectAll()

    def putlogo(self, figure: Figure):
        # ToDo: this is a repeat of mainwindow.putlogo().
        if figure is None:
            figure = self.figure
        if not hasattr(self, '_logodata'):
            self._logodata = imread(pkg_resources.resource_filename('cct', 'resource/credo_logo.png'))[:,:,0].copy()
        figure.figimage(self._logodata, 10, 10, cmap='gray', zorder=-10)
