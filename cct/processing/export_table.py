import openpyxl
import xlwt
from PyQt5 import QtCore


def export_table_xls(filename: str, model: QtCore.QAbstractItemModel):
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet1')
    assert isinstance(ws, xlwt.Worksheet)
    row = 0
    for i in range(model.columnCount()):
        ws.write(row, i, model.headerData(i, QtCore.Qt.Horizontal, QtCore.Qt.DisplayRole))
    for row in range(model.rowCount()):
        for column in range(model.columnCount()):
            value = model.index(row, column).data(QtCore.Qt.DisplayRole)
            for func in [int, float, str]:
                try:
                    value = func(value)
                    break
                except (ValueError, TypeError):
                    continue
            ws.write(row + 1, column, value)
    wb.save(filename)


def export_table_xlsx(filename: str, model: QtCore.QAbstractItemModel):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('Sheet1')
    for i in range(model.columnCount()):
        ws.cell(row=1, column=i + 1, value=model.headerData(i, QtCore.Qt.Horizontal, QtCore.Qt.DisplayRole))
    for row in range(model.rowCount()):
        for column in range(model.columnCount()):
            value = model.index(row, column).data(QtCore.Qt.DisplayRole)
            for func in [int, float, str]:
                try:
                    value = func(value)
                    break
                except (ValueError, TypeError):
                    continue
            ws.cell(column=column + 1, row=row + 2, value=value)
    wb.save(filename)


def export_table(filename: str, model: QtCore.QAbstractItemModel):
    if filename.lower().endswith('.xls'):
        return export_table_xls(filename, model)
    else:
        return export_table_xlsx(filename, model)
