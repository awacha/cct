import openpyxl
from PyQt5 import QtCore

def model2xlsx(filename:str, sheetname:str, model:QtCore.QAbstractItemModel):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheetname
    for col in range(model.columnCount()):
        ws.cell(1, col+1, value=model.headerData(col, QtCore.Qt.Horizontal, QtCore.Qt.DisplayRole))
    for row in range(model.rowCount()):
        for col in range(model.columnCount()):
            # ToDo: write numbers as numbers, dates as dates, etc, not as text.
            ws.cell(row+2, col+1, value=model.index(row, col).data(QtCore.Qt.DisplayRole))
    wb.save(filename)
