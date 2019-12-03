import logging
import os
import re
from time import monotonic
from typing import Sequence, Tuple, Optional

import numpy as np
import openpyxl
import openpyxl.utils
from PyQt5 import QtCore, QtWidgets, QtGui
from scipy.io import savemat

from .resultsdispatcher_ui import Ui_Form
from ..config import Config
from ..graphing import ImageView, CurveView, CorrMatView, VacuumFluxViewer, OutlierViewer, ExposureTimeReport, \
    TransmissionList
from ..models.results import ResultsModel
from ...qtgui.tools.anisotropy import AnisotropyEvaluator

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

toinchfactor = {'cm': 1 / 2.54,
                'mm': 0.1 / 2.54,
                'inch': 1.0}


class RegExpValidator(QtGui.QValidator):
    def validate(self, line: str, pos: int) -> Tuple[QtGui.QValidator.State, str, int]:
        try:
            re.compile(line)
            return self.Acceptable, line, pos
        except re.error:
            return self.Intermediate, line, pos
        assert False


class ResultsDispatcher(QtWidgets.QWidget, Ui_Form):
    subwindowOpenRequest = QtCore.pyqtSignal(str, QtWidgets.QWidget, name='subwindowOpenRequest')

    def __init__(self, parent: QtWidgets.QWidget, config: Config, project: "Project"):
        super().__init__(parent)
        self.config = config
        self.config.configItemChanged.connect(self.onConfigItemChanged)
        self.project = project
        self.setupUi(self)

    def setupUi(self, Form):
        super().setupUi(Form)
        self.model = ResultsModel(self.config.hdf5)
        self.treeView.setModel(self.model)
        self.model.modelReset.connect(self.resizeTreeViewColumns)
        self.model.dataChanged.connect(self.resizeTreeViewColumns)
        self.treeView.header().sortIndicatorChanged.connect(self.onTreeHeaderSortIndicatorChanged)
        self.perSampleImagePushButton.clicked.connect(self.onPerSampleImage)
        self.perSampleCurvesPushButton.clicked.connect(self.onPerSampleCurves)
        self.perSampleCorrelationMatrixPushButton.clicked.connect(self.onPersampleCorrMat)
        self.perSampleOutlierTestPushButton.clicked.connect(self.onPerSampleOutlierTest)
        self.perSampleAnisotropyPushButton.clicked.connect(self.onPerSampleAnisotropy)
        self.overallCurvesPushButton.clicked.connect(self.onOverallCurves)
        self.overallExposureTimePushButton.clicked.connect(self.onOverallExposureTime)
        self.overallTransmissionPushButton.clicked.connect(self.onOverallTransmission)
        self.overallVacuumFluxPushButton.clicked.connect(self.onOverallVacuumFlux)
        self.selectAllSamplesToolButton.clicked.connect(self.selectAllSamples)
        self.selectNoSamplesToolButton.clicked.connect(self.selectNoSamples)
        self.selectRegexToolButton.clicked.connect(self.selectRegex)
        self.deselectRegexToolButton.clicked.connect(self.deselectRegex)
        self.regexpValidator = RegExpValidator()
        self.sampleNameRegexLineEdit.setValidator(self.regexpValidator)
        self.sampleNameRegexLineEdit.textChanged.connect(self.onRegexInvalid)
        self.curveFileFormatComboBox.addItems(self.config.acceptableValues('onedimformat'))
        self.curveFileFormatComboBox.setCurrentIndex(self.curveFileFormatComboBox.findText(self.config.onedimformat))
        self.patternFileFormatComboBox.addItems(self.config.acceptableValues('twodimformat'))
        self.patternFileFormatComboBox.setCurrentIndex(
            self.patternFileFormatComboBox.findText(self.config.twodimformat))
        self.cmatFileFormatComboBox.addItems(self.config.acceptableValues('twodimformat'))
        self.cmatFileFormatComboBox.setCurrentIndex(self.cmatFileFormatComboBox.findText(self.config.twodimformat))
        self.exportDirLineEdit.setText(self.project.config.folder)
        self.exportCurvesPushButton.clicked.connect(self.exportCurves)
        self.exportPatternsPushButton.clicked.connect(self.exportPatterns)
        self.exportCorrelMatricesPushButton.clicked.connect(self.exportCorrelMatrices)
        self.exportDirToolButton.clicked.connect(self.browseExportFolder)
        self.exportDirLineEdit.textChanged.connect(self.exportFolderChanged)
        self.exportDirLineEdit.editingFinished.connect(self.exportFolderEditingFinished)
        self.graphResolutionSpinBox.setValue(self.config.imagedpi)
        self.graphHeightDoubleSpinBox.setValue(self.config.imagewidth / toinchfactor[self.config.imagewidthunits])
        self.graphWidthDoubleSpinBox.setValue(self.config.imageheight / toinchfactor[self.config.imageheightunits])
        self.graphWidthDoubleSpinBox.setSuffix(' {}'.format(self.config.imagewidthunits))
        self.graphHeightDoubleSpinBox.setSuffix(' {}'.format(self.config.imageheightunits))
        self.graphFormatComboBox.addItems(sorted(self.config.acceptableValues('imageformat')))
        self.graphFormatComboBox.setCurrentIndex((self.graphFormatComboBox.findText(self.config.imageformat)))
        self.exportCorrelMatricesGraphPushButton.clicked.connect(self.exportCorrelMatricesGraph)
        self.exportCurvesGraphPushButton.clicked.connect(self.exportCurvesGraph)
        self.exportPatternsGraphPushButton.clicked.connect(self.exportPatternsGraph)
        self.exportProgressBar.setVisible(False)
        self.treeView.selectionModel().selectionChanged.connect(self.updateCommandWidgetsSensitivity)
        self.updateCommandWidgetsSensitivity()
        self.resizeTreeViewColumns()

    def onRegexInvalid(self):
        self.selectRegexToolButton.setEnabled(self.sampleNameRegexLineEdit.hasAcceptableInput())
        self.deselectRegexToolButton.setEnabled(self.sampleNameRegexLineEdit.hasAcceptableInput())

    def selectAllSamples(self):
        self.treeView.selectAll()

    def selectNoSamples(self):
        self.treeView.clearSelection()

    def updateCommandWidgetsSensitivity(self):
        selectedCount = self.countSelectedResults()
        notSubtractedSelectedCount = self.countSelectedResults(subtractedtoo=False)
        hasexportDir = os.path.isdir(self.exportDirLineEdit.text())
        for widget in [self.perSampleAnisotropyPushButton,
                       self.perSampleImagePushButton,
                       self.overallCurvesPushButton]:
            # these widgets need only selected samples
            widget.setEnabled(selectedCount > 0)
        for widget in [self.exportCurvesGraphPushButton, self.exportPatternsGraphPushButton,
                       self.exportCorrelMatricesGraphPushButton,
                       self.exportCurvesPushButton, self.exportPatternsPushButton, self.exportCorrelMatricesPushButton]:
            # these widgets need selected samples AND a valid output dir
            widget.setEnabled((selectedCount > 0) and hasexportDir)
        for widget in [self.perSampleOutlierTestPushButton, self.perSampleCorrelationMatrixPushButton,
                       self.perSampleCurvesPushButton, self.overallVacuumFluxPushButton,
                       self.overallExposureTimePushButton, self.overallTransmissionPushButton]:
            # these widgets need selected samples which are not subtracted samples
            widget.setEnabled(notSubtractedSelectedCount > 0)

    def selectRegex(self):
        try:
            regexp = re.compile(self.sampleNameRegexLineEdit.text())
        except re.error:
            assert False
        for i in range(self.model.rowCount()):
            if regexp.match(self.model[i].samplename):
                self.treeView.selectionModel().select(self.model.index(i, 0),
                                                      QtCore.QItemSelectionModel.Select | QtCore.QItemSelectionModel.Rows)

    def deselectRegex(self):
        try:
            regexp = re.compile(self.sampleNameRegexLineEdit.text())
        except re.error:
            assert False
        for i in range(self.model.rowCount()):
            if regexp.match(self.model[i].samplename):
                self.treeView.selectionModel().select(self.model.index(i, 0),
                                                      QtCore.QItemSelectionModel.Deselect | QtCore.QItemSelectionModel.Rows)

    def onOverallCurves(self):
        cv = CurveView(None, self.project)
        for sample, distance in self.selectedResults():
            logger.debug('Adding curve to overall curve comparison: {}, {}'.format(sample, distance))
            curve = self.project.h5reader.averagedCurve(sample, distance)
            cv.addCurve(curve, (sample, distance), label='{} @{:.2f} mm'.format(sample, distance))
        cv.replot()
        cv.setWindowTitle('Compare curves')
        self.subwindowOpenRequest.emit('curve_compare_{}'.format(monotonic()), cv)

    def onOverallExposureTime(self):
        etr = ExposureTimeReport(None, self.project)
        for samplename, distance in self.selectedResults():
            try:
                self.project.h5reader.getCurveParameter(samplename, distance, 'fsn')
            except KeyError:
                continue
            etr.addSampleAndDistance(samplename, distance, False)
        etr.replot()
        self.subwindowOpenRequest.emit('exptimereport_{}'.format(monotonic()), etr)

    def onOverallTransmission(self):
        tv = TransmissionList(None, self.project)
        for samplename, distance in self.selectedResults():
            tv.addSampleAndDist(samplename, distance, updatelist=False)
        tv.updateList()
        self.subwindowOpenRequest.emit('transmission_{}'.format(monotonic()), tv)

    def onOverallVacuumFlux(self):
        vf = VacuumFluxViewer(None, self.project)
        for samplename, distance in self.selectedResults():
            try:
                self.project.h5reader.getCurveParameter(samplename, distance, 'fsn')
            except KeyError:
                continue
            vf.addSampleAndDist(samplename, distance, replot=False)
        vf.replot()
        self.subwindowOpenRequest.emit('vacuumandflux_{}'.format(monotonic()), vf)

    def onTreeHeaderSortIndicatorChanged(self, logicalsection: int, sortorder: QtCore.Qt.SortOrder):
        logger.debug('Sort order changed: section {}, order {}'.format(logicalsection, sortorder))
        self.treeView.sortByColumn(logicalsection, sortorder)

    def resizeTreeViewColumns(self):
        for c in range(self.model.columnCount()):
            self.treeView.resizeColumnToContents(c)

    def reloadResults(self):
        self.model.reload()
        self.model.sort(self.treeView.header().sortIndicatorSection(), self.treeView.header().sortIndicatorOrder())

    def onConfigItemChanged(self, section, itemname, newvalue):
        if itemname == 'hdf5':
            self.model.setH5FileName(newvalue)
        elif itemname == 'imagewidthunits':
            self.updateImageWidthUnits()
        elif itemname == 'imageheightunits':
            self.updateImageHeightUnits()
        elif itemname == 'folder':
            self.exportDirLineEdit.setText(newvalue)

    def updateImageWidthUnits(self):
        prevunits = self.graphWidthDoubleSpinBox.suffix().strip()
        if prevunits:
            value_in_inch = toinchfactor[prevunits] * self.graphWidthDoubleSpinBox.value()
            self.graphWidthDoubleSpinBox.setValue(value_in_inch / toinchfactor[self.config.imagewidthunits])
        self.graphWidthDoubleSpinBox.setSuffix(' ' + self.config.imagewidthunits)

    def updateImageHeightUnits(self):
        prevunits = self.graphHeightDoubleSpinBox.suffix().strip()
        if prevunits:
            value_in_inch = toinchfactor[prevunits] * self.graphHeightDoubleSpinBox.value()
            self.graphHeightDoubleSpinBox.setValue(value_in_inch / toinchfactor[self.config.imageheightunits])
        self.graphHeightDoubleSpinBox.setSuffix(' ' + self.config.imageheightunits)

    def selectedResults(self) -> Sequence[Tuple[str, float]]:
        for idx in self.treeView.selectionModel().selectedRows(0):
            yield (self.model[idx.row()].samplename, float(self.model[idx.row()].distance))
        return

    def countSelectedResults(self, subtractedtoo: bool = True) -> int:
        return len([index for index in self.treeView.selectionModel().selectedRows(0)
                    if subtractedtoo or (self.model[index.row()].samplecategory != 'subtracted')])

    def onPerSampleImage(self):
        for samplename, distance in self.selectedResults():
            iv = ImageView(None, config=self.project.config)
            iv.setExposure(self.project.h5reader.averagedImage(samplename, distance))
            iv.setWindowTitle('{} @ {:.2f}'.format(samplename, distance))
            self.subwindowOpenRequest.emit('persampleImage_{}'.format(monotonic()), iv)

    def onPerSampleCurves(self):
        for samplename, distance in self.selectedResults():
            try:
                allcurves = self.project.h5reader.allCurves(samplename, distance)
            except KeyError:
                continue
            cv = CurveView(None, self.project)
            avg = self.project.h5reader.averagedCurve(samplename, distance)
            badfsns = self.project.h5reader.badFSNs(samplename, distance)
            for fsn, curve in allcurves.items():
                cv.addCurve(curve, color='red' if int(fsn) in badfsns else 'green', label='#{}'.format(int(fsn)))
            cv.addCurve(avg, color='black', lw=1, label='Mean')
            cv.replot()
            cv.setWindowTitle('All exposures of {} @ {}'.format(samplename, distance))
            self.subwindowOpenRequest.emit('persampleCurve_{}_{}_{}'.format(samplename, distance, monotonic()), cv)

    def onPersampleCorrMat(self):
        for samplename, distance in self.selectedResults():
            try:
                cmat = self.project.h5reader.getCorrMat(samplename, distance)
                fsns = sorted(list(self.project.h5reader.getCurveParameter(samplename, distance, 'fsn').keys()))
            except KeyError:
                continue
            cmatview = CorrMatView(None, self.project)
            cmatview.setSampleAndDistance(samplename, distance)
            self.subwindowOpenRequest.emit('cmat_{}_{}_{}'.format(samplename, distance, monotonic()), cmatview)

    def onPerSampleOutlierTest(self):
        for samplename, distance in self.selectedResults():
            try:
                self.project.h5reader.getCorrMat(samplename, distance)
            except KeyError:
                continue
            ov = OutlierViewer(None, self.project)
            ov.setSampleAndDistance(samplename, distance)
            self.subwindowOpenRequest.emit('outlierviewer_{}_{}_{}'.format(samplename, distance, monotonic()), ov)

    def onPerSampleAnisotropy(self):
        for samplename, distance in self.selectedResults():
            ae = AnisotropyEvaluator()
            ae.h5Selector.close()
            ae.h5Selector.destroy()
            ae.setExposure(self.project.h5reader.averagedImage(samplename, distance))
            self.subwindowOpenRequest.emit('persampleAnisotropy_{}_{}_{}'.format(samplename, distance, monotonic()), ae)

    def exportCurves(self):
        self._startExport('Exporting curves...')
        if self.curveFileFormatComboBox.currentText() == 'ASCII (*.txt)':
            for i, (samplename, distance) in enumerate(self.selectedResults()):
                self.exportProgress(i)
                try:
                    curve = self.project.h5reader.averagedCurve(samplename, distance)
                except KeyError:
                    continue  # ToDo: present an error message
                with open(os.path.join(self.exportDirLineEdit.text(),
                                       '{}_{:.2f}.txt'.format(samplename, float(distance))), 'wt') as f:
                    dic = self.project.h5reader.averagedHeaderDict(samplename, distance)
                    for key in sorted(dic):
                        f.write('# {} : {}\n'.format(key, dic[key]))
                    f.write('# Columns:\n')
                    f.write(
                        '#  q [1/nm],  Intensity [1/cm * 1/sr], '
                        'Propagated uncertainty of the intensity [1/cm * 1/sr], Propagated uncertainty of q [1/nm]\n')
                    np.savetxt(f, np.vstack((curve.q, curve.Intensity, curve.Error, curve.qError)).T)
        elif self.curveFileFormatComboBox.currentText() == 'ASCII (*.dat)':
            for i, (samplename, distance) in enumerate(self.selectedResults()):
                self.exportProgress(i)
                try:
                    curve = self.project.h5reader.averagedCurve(samplename, distance)
                except KeyError:
                    continue  # ToDo: present an error message
                with open(os.path.join(self.exportDirLineEdit.text(),
                                       '{}_{:.2f}.dat'.format(samplename, float(distance))), 'wt') as f:
                    dic = self.project.h5reader.averagedHeaderDict(samplename, distance)
                    for key in sorted(dic):
                        f.write('# {} : {}\n'.format(key, dic[key]))
                    f.write('# Columns:\n')
                    f.write(
                        '#  q [1/nm],  Intensity [1/cm * 1/sr], '
                        'Propagated uncertainty of the intensity [1/cm * 1/sr]\n')
                    np.savetxt(f, np.vstack((curve.q, curve.Intensity, curve.Error)).T)
        elif self.curveFileFormatComboBox.currentText() == 'ATSAS (*.dat)':
            for i, (samplename, distance) in enumerate(self.selectedResults()):
                self.exportProgress(i)
                try:
                    curve = self.project.h5reader.averagedCurve(samplename, distance)
                except KeyError:
                    continue  # ToDo: present an error message
                with open(os.path.join(self.exportDirLineEdit.text(),
                                       '{}_{:.2f}.dat'.format(samplename, float(distance))), 'wt') as f:
                    dic = self.project.h5reader.averagedHeaderDict(samplename, distance)
                    for key in sorted(dic):
                        f.write('# {} : {}\n'.format(key, dic[key]))
                    f.write('# Columns:\n')
                    f.write(
                        '#  q [1/A],  Intensity [1/cm * 1/sr], Propagated uncertainty of the intensity [1/cm * 1/sr]\n')
                    np.savetxt(f, np.vstack((curve.q / 10, curve.Intensity, curve.Error)).T)
        elif self.curveFileFormatComboBox.currentText() == 'RSR (*.rsr)':
            for i, (samplename, distance) in enumerate(self.selectedResults()):
                self.exportProgress(i)
                try:
                    curve = self.project.h5reader.averagedCurve(samplename, distance)
                except KeyError:
                    continue  # ToDo: present an error message
                with open(os.path.join(self.exportDirLineEdit.text(),
                                       '{}_{:.2f}.rsr'.format(samplename, float(distance))), 'wt') as f:
                    f.write(' TIME\n')
                    f.write(' 1.0\n')
                    f.write(' {}\n'.format(len(curve)))
                    for j in range(len(curve)):
                        f.write(
                            ' {:.9f} {:.9f} {:.9f} 1\n'.format(curve.q[j] / 10., curve.Intensity[j], curve.Error[j]))
        elif self.curveFileFormatComboBox.currentText() == 'PDH (*.pdh)':
            for i, (samplename, distance) in enumerate(self.selectedResults()):
                self.exportProgress(i)
                try:
                    curve = self.project.h5reader.averagedCurve(samplename, distance)
                except KeyError:
                    continue  # ToDo: present an error message
                with open(os.path.join(self.exportDirLineEdit.text(),
                                       '{}_{:.2f}.pdh'.format(samplename, float(distance))), 'wt') as f:
                    f.write('{} @ {:.2f}\n'.format(samplename, float(distance)))
                    f.write('SAXS\n')
                    f.write(
                        '{:>9d} {:>9d} {:>9d} {:>9d} {:>9d} {:>9d} {:>9d} {:>9d}\n'.format(len(curve), 0, 0, 0, 0, 0, 0,
                                                                                           0))
                    f.write('{:>14.6E} {:>14.6E} {:>14.6E} {:>14.6E} {:>14.6E}\n'.format(0, 0, 0, 1, 0))
                    f.write('{:>14.6E} {:>14.6E} {:>14.6E} {:>14.6E} {:>14.6E}\n'.format(0, 0, 0, 0, 0))
                    for j in range(len(curve)):
                        f.write('{:>14.6E} {:>14.6E} {:>14.6E}\n'.format(curve.q[j] / 10, curve.Intensity[j],
                                                                         curve.Error[j]))
        elif self.curveFileFormatComboBox.currentText() == 'Excel 2007- (*.xlsx)':
            wb = openpyxl.Workbook()
            ws_main = wb.active
            ws_main.title = 'Summary'
            sheetindex = 0
            for i, (samplename, distance) in enumerate(self.selectedResults()):
                self.exportProgress(i)
                try:
                    curve = self.project.h5reader.averagedCurve(samplename, distance)
                except KeyError:
                    continue  # ToDo: present an error message
                sheetname = '{}_{:.2f}'.format(samplename, float(distance))
                ws = wb.create_sheet(sheetname)
                ws_main.cell(row=1, column=4 * sheetindex + 1, value=sheetname)
                ws_main.merge_cells(start_row=1, start_column=4 * sheetindex + 1, end_row=1,
                                    end_column=4 * sheetindex + 4)
                ws.cell(row=1, column=1, value='q')
                ws.cell(row=1, column=2, value='Intensity')
                ws.cell(row=1, column=3, value='dIntensity')
                ws.cell(row=1, column=4, value='dq')
                ws.cell(row=2, column=1, value='1/nm')
                ws.cell(row=2, column=2, value='1/cm * 1/sr')
                ws.cell(row=2, column=3, value='1/cm * 1/sr')
                ws.cell(row=2, column=4, value='1/nm')
                for j in range(len(curve)):
                    ws.cell(row=3 + j, column=1, value=curve.q[j])
                    ws.cell(row=3 + j, column=2, value=curve.Intensity[j])
                    ws.cell(row=3 + j, column=3, value=curve.Error[j])
                    ws.cell(row=3 + j, column=4, value=curve.qError[j])
                for row in range(1, 3 + len(curve) + 1):
                    for column in range(1, 5):
                        ws_main.cell(row=row + 1, column=4 * sheetindex + column, value='={}!{}{}'.format(
                            openpyxl.utils.quote_sheetname(sheetname), openpyxl.utils.get_column_letter(column),
                            row
                        ))
                sheetindex += 1
            wb.save(os.path.join(self.exportDirLineEdit.text(), 'SAXS_curves.xlsx'))
        else:
            raise ValueError('Unknown 1D file format: {}'.format(self.curveFileFormatComboBox.currentText()))
        self._finishExport()

    def exportPatterns(self):
        self._startExport('Exporting patterns...')
        for i, (samplename, distance) in enumerate(self.selectedResults()):
            self.exportProgress(i)
            try:
                exposure = self.project.h5reader.averagedImage(samplename, distance)
            except KeyError:
                continue  # ToDo: warn the user.
            if self.patternFileFormatComboBox.currentText() == 'Numpy (*.npz)':
                np.savez_compressed(
                    os.path.join(self.exportDirLineEdit.text(), '{}_{:.2f}.npz'.format(samplename, float(distance))),
                    intensity=exposure.intensity,
                    error=exposure.error,
                    mask=exposure.mask
                )
            elif self.patternFileFormatComboBox.currentText() == 'Matlab(TM) (*.mat)':
                savemat(
                    os.path.join(self.exportDirLineEdit.text(), '{}_{:.2f}.mat'.format(samplename, float(distance))),
                    {'intensity': exposure.intensity,
                     'error': exposure.error,
                     'mask': exposure.mask},
                    do_compression=True
                )
            elif self.patternFileFormatComboBox.currentText() in ['Gzip-ped ASCII (*.txt.gz)', 'ASCII (*.txt)']:
                extn = '.txt.gz' if self.patternFileFormatComboBox.currentText().lower().startswith('gzip') else '.txt'
                basename = os.path.join(self.exportDirLineEdit.text(), '{}_{:.2f}'.format(samplename, float(distance)))
                np.savetxt(basename + '_intensity' + extn, exposure.intensity)
                np.savetxt(basename + '_error' + extn, exposure.error)
                np.savetxt(basename + '_mask' + extn, exposure.mask)
            else:
                raise ValueError('Unknown file format: {}'.format(self.patternFileFormatComboBox.currentText()))
        self._finishExport()

    def exportCorrelMatrices(self):
        self._startExport('Exporting correlation matrices...')
        for i, (samplename, distance) in enumerate(self.selectedResults()):
            self.exportProgress(i)
            try:
                cmat = self.project.h5reader.getCorrMat(samplename, distance)
            except KeyError:
                continue  # ToDo: warn the user.
            if self.patternFileFormatComboBox.currentText() == 'Numpy (*.npz)':
                np.savez_compressed(
                    os.path.join(self.exportDirLineEdit.text(),
                                 'correlmatrix_{}_{:.2f}.npz'.format(samplename, float(distance))),
                    correlmatrix=cmat)
            elif self.patternFileFormatComboBox.currentText() == 'Matlab(TM) (*.mat)':
                savemat(
                    os.path.join(self.exportDirLineEdit.text(),
                                 'correlmatrix_{}_{:.2f}.mat'.format(samplename, float(distance))),
                    {'correlmatrix': cmat},
                    do_compression=True
                )
            elif self.patternFileFormatComboBox.currentText() in ['Gzip-ped ASCII (*.txt.gz)', 'ASCII (*.txt)']:
                extn = '.txt.gz' if self.patternFileFormatComboBox.currentText().lower().startswith('gzip') else '.txt'
                basename = os.path.join(self.exportDirLineEdit.text(),
                                        'correlmatrix_{}_{:.2f}'.format(samplename, float(distance)))
                np.savetxt(basename + extn, cmat)
            else:
                raise ValueError('Unknown file format: {}'.format(self.patternFileFormatComboBox.currentText()))
        self._finishExport()

    def browseExportFolder(self):
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, 'Export files to...', '')
        if not folder:
            return
        self.exportDirLineEdit.setText(folder)
        self.exportFolderEditingFinished()

    def exportFolderChanged(self):
        logger.debug('Export folder changed to: {}'.format(self.exportDirLineEdit.text()))
        self.updateCommandWidgetsSensitivity()

    def exportFolderEditingFinished(self):
        logger.debug('Export folder edited to: {}'.format(self.exportDirLineEdit.text()))
        self.updateCommandWidgetsSensitivity()
        self.project.config.folder = self.exportDirLineEdit.text()

    def exportCurvesGraph(self):
        # drawing the graph is tricky: we create a figure and a parent-less canvas, which we will never show and will
        # destroy when the graph has been saved
        self._startExport('Drawing curves...')
        cv = CurveView(None, self.project, figsize=(self.graphWidthDoubleSpinBox.value() * toinchfactor[
            self.graphWidthDoubleSpinBox.suffix().strip()],
                                                    self.graphHeightDoubleSpinBox.value() * toinchfactor[
                                                        self.graphHeightDoubleSpinBox.suffix().strip()]))
        for i, (sample, distance) in enumerate(self.selectedResults()):
            self.exportProgress(i)
            curve = self.project.h5reader.averagedCurve(sample, distance)
            cv.addCurve(curve, label='{}_{:.2f}'.format(sample, float(distance)))
        cv.replot()
        cv.savefig(os.path.join(self.exportDirLineEdit.text(), 'saxs_curves.' + self.graphFormatComboBox.currentText()),
                   dpi=self.graphResolutionSpinBox.value())
        cv.close()
        cv.destroy()
        cv.deleteLater()
        self._finishExport()

    def exportPatternsGraph(self):
        self._startExport('Drawing patterns...')
        iv = ImageView(None, self.project.config,
                       figsize=(self.graphWidthDoubleSpinBox.value() * toinchfactor[
                           self.graphWidthDoubleSpinBox.suffix().strip()],
                                self.graphHeightDoubleSpinBox.value() * toinchfactor[
                                    self.graphHeightDoubleSpinBox.suffix().strip()]))
        for i, (sample, distance) in enumerate(self.selectedResults()):
            self.exportProgress(i)
            ex = self.project.h5reader.averagedImage(sample, distance)
            iv.setExposure(ex)
            iv.savefig(os.path.join(self.exportDirLineEdit.text(), '{}_{:.2f}.{}'.format(
                sample, float(distance),
                self.graphFormatComboBox.currentText())))
        iv.close()
        iv.destroy()
        iv.deleteLater()
        self._finishExport()

    def exportCorrelMatricesGraph(self):
        self._startExport('Drawing correlation matrices...')
        cv = CorrMatView(None, self.project, figsize=(self.graphWidthDoubleSpinBox.value() * toinchfactor[
            self.graphWidthDoubleSpinBox.suffix().strip()],
                                                      self.graphHeightDoubleSpinBox.value() * toinchfactor[
                                                          self.graphHeightDoubleSpinBox.suffix().strip()]))
        for i, (sample, distance) in enumerate(self.selectedResults()):
            self.exportProgress(i)
            try:
                self.project.h5reader.getCorrMat(sample, distance)
            except KeyError:
                continue
            cv.setSampleAndDistance(sample, distance)
            cv.savefig(os.path.join(self.exportDirLineEdit.text(), 'corrmat_{}_{:.2f}.{}'.format(
                sample, float(distance),
                self.graphFormatComboBox.currentText())))
        cv.close()
        cv.destroy()
        cv.deleteLater()
        self._finishExport()

    def _startExport(self, message: str = 'Exporting...'):
        self.setEnabled(False)
        self.exportProgressBar.setMinimum(0)
        self.exportProgressBar.setMaximum(self.countSelectedResults())
        self.exportProgressBar.setFormat(message)
        self.exportProgressBar.setEnabled(True)
        self.exportProgressBar.setVisible(True)
        self.exportProgressBar.setValue(0)

    def _finishExport(self):
        self.setEnabled(True)
        self.exportProgressBar.setVisible(False)

    def exportProgress(self, current: int, message: Optional[str] = None):
        self.exportProgressBar.setValue(current)
        if message is not None:
            self.exportProgressBar.setFormat(message)
        QtWidgets.QApplication.instance().processEvents()
        QtWidgets.QApplication.instance().sendPostedEvents()
