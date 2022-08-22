import logging
import os
from typing import Final, List

import openpyxl
import openpyxl.styles
import openpyxl.utils

from ..core2.config import Config
from ..core2.dataclasses import Header

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

MAXNOTFOUNDCOUNT: Final[int] = 100


def listsubdirs(path: str):
    yield path
    for fn in os.listdir(path):
        if os.path.isdir(fn):
            yield from listsubdirs(os.path.join(path, fn))


columns: Final[List[str]] = ['fsn', 'title', 'distance', 'distancedecrease', 'exposuretime', 'temperature',
                             'thickness', 'transmission', 'date', 'beamposrow', 'beamposcol', 'samplex', 'sampley',
                             'maskname', 'vacuum', 'username', 'project', 'fsn_emptybeam', 'fsn_absintref', 'fsn_dark',
                             'absintfactor', 'flux']


def mkexcel(filename: str, firstfsn: int, lastfsn: int, configfile: str, verbose: bool):
    logger.setLevel(logging.INFO if verbose else logging.INFO)
    config = Config(dicorfile=configfile)
    config.filename = None  # inhibit auto-save
    wb = openpyxl.Workbook()
    ws = wb.active
    bold = openpyxl.styles.Font(bold=True)
    align = openpyxl.styles.Alignment(horizontal='center')
    for col, colname in enumerate(columns, start=1):
        ws.cell(row=1, column=col, value=colname.capitalize()).font = bold
        ws.cell(row=1, column=col).alignment = align
    subdirs = []
    for subpath in ['eval2d', 'param_override', 'param']:
        subdirs.extend(listsubdirs(config['path']['directories'][subpath]))
    row = 2
    for fsn in range(firstfsn, lastfsn + 1):
        for subdir in subdirs:
            try:
                header = Header(
                    filename=os.path.join(
                        subdir,
                        f'{config["path"]["prefixes"]["crd"]}_{fsn:0{config["path"]["fsndigits"]}d}.pickle'))
                logger.debug(f'Found header {header.fsn} in {subdir}')
                for i, col in enumerate(columns, start=1):
                    try:
                        value = getattr(header, col)
                    except (KeyError, AttributeError):
                        value = None
                    if isinstance(value, tuple):
                        value = value[0]
                    ws.cell(row=row, column=i, value=value)
                row += 1
                break  # do not look further in other subdirectories
            except FileNotFoundError:
                # file not found in this subdirectory: look for it in another one.
                continue
    ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(columns))}{row}'
    wb.save(filename)
